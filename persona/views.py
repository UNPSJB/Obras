# coding=utf-8
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login
from django.contrib.auth.decorators import  login_required,user_passes_test
import easygui as eg
from .forms import *
from django.contrib import messages
from pago.forms import *
from tipos.forms import *
from obras_particulares.views import *
from tramite.forms import FormularioIniciarTramite
from documento.forms import FormularioDocumentoSetFactory
from documento.forms import metodo
from tramite.models import *
from django.core.mail import send_mail
from persona.models import *
from tramite.models import Tramite, Estado
from django.views.generic.detail import DetailView
import re
from django.views.generic.base import TemplateView
from openpyxl import Workbook
from django.http.response import HttpResponse
from django.views.generic import View
from django.conf import settings
from io import BytesIO
from datetime import date

from reportlab.pdfgen import canvas
from reportlab.platypus import SimpleDocTemplate, Paragraph, TableStyle, Table, Image, Spacer
from reportlab.lib import colors
from reportlab.lib.units import cm, inch
from reportlab.lib.pagesizes import letter, A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_JUSTIFY, TA_LEFT, TA_RIGHT
from reportlab.lib.validators import Auto
from reportlab.graphics.charts.legends import Legend
from reportlab.graphics.charts.piecharts import Pie
from reportlab.graphics.shapes import Drawing, String
import base64;
import time
import collections
from planilla_visado.models import ItemDeVisado
from pago.models import Cuota, Cancelacion,Cancelada,Estado
from datetime import datetime, date, time, timedelta
from documento.models import Documento
from reportlab.graphics.charts.linecharts import HorizontalLineChart
from reportlab.lib.colors import Color, PCMYKColor, white
from reportlab.graphics.charts.barcharts import VerticalBarChart
from calendar import monthrange
from django.db.models import Max
from django.db.models import Count
from collections import defaultdict
import json
from itertools import chain
import math
#-------------------------------------------------------------------------------------------------------------------
#generales ---------------------------------------------------------------------------------------------------------


DATETIME = re.compile("^(\d{4})\-(\d{2})\-(\d{2})\s(\d{2}):(\d{2})$")

def convertidor_de_fechas(fecha):
    return datetime.datetime(*[int(n) for n in DATETIME.match(fecha).groups()])    

#-------------------------------------------------------------------------------------------------------------------
#propietario -------------------------------------------------------------------------------------------------------

@login_required(login_url="login")
@grupo_requerido('propietario')
def mostrar_propietario(request):
    estilos = ''
    usuario = request.user
    propietario = get_object_or_404(Propietario, pk=usuario.persona.propietario.pk)
    if request.method == "POST":
        if "Aceptar" in request.POST:
            estilos = request.POST.get('estilo')
            propietario = Propietario.objects.filter(id=usuario.persona.propietario.pk).update(estilo=estilos)
    else:
        if propietario.estilo:
            estilos = propietario.estilo
    contexto = {
        "ctxtramitespropietario": tramites_de_propietario(request),
        "ctxmis_tramites_para_financiar": tramites_para_financiar(request),
        "ctxtramites_para_financiar_propietario": listado_tramites_para_financiar_propietario(request),
        "ctxlistado": listado_tramites_propietario(request),
        'estilos': estilos,
    }
    return render(request, 'persona/propietario/propietario.html', contexto)

def tramites_para_financiar(request):
    usuario = request.user
    persona = Persona.objects.get(usuario__isnull=False, usuario_id=usuario)
   # propietario = persona.get_propietario()  # Me quedo con el atributo propietario de la persona
    tramites_propietario = Tramite.objects.en_estado(Visado)
    tramites = filter(lambda tramite: (tramite.propietario == persona.propietario and tramite.pago is None), tramites_propietario)
    return tramites

def tramites_de_propietario(request):
    tramites = Tramite.objects.all()
    usuario = request.user
    persona = Persona.objects.get(usuario__isnull=False, usuario_id=usuario)
    #propietario = persona.get_propietario()  # Me quedo con el atributo propietario de la persona
    #tramites_de_propietario = filter(lambda tramite: (tramite.propietario == persona.propietario and  tramite.pago_id is not None), tramites)
    tramites_de_propietario = filter(
        lambda tramite: (tramite.propietario == persona.propietario), tramites)
    return tramites_de_propietario

def propietario_solicita_final_obra(request, pk_tramite):
    tramite = get_object_or_404(Tramite, pk=pk_tramite)
    try:
        if tramite.estados.filter(tipo=8):
            messages.add_message(request, messages.WARNING,"La solicitud de final de obra ya fue otorgada anteriormente")
        else:
            tramite.hacer(Tramite.SOLICITAR_FINAL_OBRA, request.user)
            messages.add_message(request, messages.SUCCESS, 'final de obra solicitado.')
    except:
        messages.add_message(request, messages.ERROR, 'No puede solicitar el final de obra para ese tramite.')
    finally:
        return redirect('propietario')

def ver_historial_tramite(request, pk_tramite):
    estilos = ''
    usuario = request.user
    propietario = get_object_or_404(Propietario, pk=usuario.persona.propietario.pk)
    if propietario.estilo:
        estilos = propietario.estilo
    tramite = get_object_or_404(Tramite, pk=pk_tramite)
    contexto0 = {'tramite': tramite}
    estados_de_tramite = tramite.estadosTramite()
    contexto1 = {'estados_del_tramite': estados_de_tramite}
    fechas_del_estado = [];
    for est in estados_de_tramite:
        fechas_del_estado.append(est.timestamp.strftime("%d/%m/%Y"));
    return render(request, 'persona/propietario/ver_historial_tramite.html', {"tramite": contexto0, "estadosp": contexto1, "fecha":fechas_del_estado,'estilos':estilos})

def documentos_de_estado(request, pk_estado):
    estilos = ''
    usuario = request.user
    propietario = get_object_or_404(Propietario, pk=usuario.persona.propietario.pk)
    if propietario.estilo:
        estilos = propietario.estilo
    estado = get_object_or_404(Estado, pk=pk_estado)
    return render(request, 'persona/propietario/documentos_de_estado.html', {'documentos':estado.tramite.documentacion_para_estado(estado),'estilos':estilos})
'''
def documentos_de_estado(request, pk_estado):
    estilos = ''
    usuario = request.user
    propietario = get_object_or_404(Propietario, pk=usuario.persona.propietario.pk)
    if propietario.estilo:
        estilos = propietario.estilo
    estado = get_object_or_404(Estado, pk=pk_estado)
    fecha = estado.timestamp
    fecha_str = date.strftime(fecha, '%d/%m/%Y %H:%M')
    documentos = estado.tramite.documentacion_para_estado(estado)
    documentosF = estado.tramite.documentos.all()
    documentos = filter(lambda e:(date.strftime(e.fecha, '%d/%m/%Y %H:%M') <= fecha_str), documentosF)
    planillas = []
    inspecciones = []
    if (estado.tipo == 1 or estado.tipo == 2):
        contexto = {'documentos': documentos, 'estilos': estilos}
    if (estado.tipo >2 and estado.tipo < 5):
        for p in PlanillaDeVisado.objects.all():
            if (p.tramite.pk == estado.tramite.pk):
                planillas.append(p)
        filas = FilaDeVisado.objects.all()
        columnas = ColumnaDeVisado.objects.all()
        contexto = {
            'planillas':planillas,
            'filas': filas,
            'columnas': columnas,
            'estilos':estilos
        }
    if (estado.tipo >= 5 and estado.tipo <= 9):
        for p in PlanillaDeInspeccion.objects.all():
            if (p.tramite.pk == estado.tramite.pk):
                inspecciones.append(p)
        items = ItemInspeccion.objects.all()
        categorias = CategoriaInspeccion.objects.all()
        contexto = {
            'inspecciones': inspecciones,
            'items': items,
            'categorias': categorias,
            'estilos':estilos
        }
    return render(request, 'persona/propietario/documentos_de_estado.html', contexto)
'''
def listado_tramites_para_financiar_propietario(request):
    estilos = ''
    usuario = request.user
    propietario = get_object_or_404(Propietario, pk=usuario.persona.propietario.pk)
    if propietario.estilo:
        estilos = propietario.estilo
    tramites = Tramite.objects.en_estado([Visado, Agendado, ConInspeccion, Inspeccionado, FinalObraSolicitado])
    usuario = request.user
    persona = Persona.objects.get(usuario__isnull=False, usuario_id=usuario)
    tramites_de_propietario = filter(lambda tramite: (tramite.propietario == persona.propietario and tramite.pago_id is None), tramites)
    contexto = {'tramites':tramites_de_propietario, 'estilos':estilos}
    return contexto

def listado_de_comprobantes_propietario(request, pk_tramite):
    estilos = ''
    usuario = request.user
    propietario = get_object_or_404(Propietario, pk=usuario.persona.propietario.pk)
    if propietario.estilo:
        estilos = propietario.estilo
    tramite = get_object_or_404(Tramite, pk=pk_tramite)
    tramites = Tramite.objects.filter(propietario = propietario)
    list = []
    for t in tramites:
        list.append(t.pago_id)
        pago = tramite.pago
        canceladas = []
        cuotas = Cuota.objects.en_estado(Cancelada)
        for cuota in cuotas:
            if cuota.pago == pago:
                canceladas.append(cuota)
        if canceladas is None:
            messages.add_message(request, messages.WARNING, 'No hay pagos registrados para el tramite seleccionado.')
        return render(request, 'persona/propietario/listado_de_comprobantes_propietario.html', {'cuotas': canceladas,'tramite':tramite, 'estilos':estilos})

def elegir_financiacion_propietario(request,pk_tramite):
    estilos = ''
    usuario = request.user
    propietario = get_object_or_404(Propietario, pk=usuario.persona.propietario.pk)
    if propietario.estilo:
        estilos = propietario.estilo
    tramite = get_object_or_404(Tramite, pk=pk_tramite)
    if request.method == "POST":
        if "Guardar" in request.POST:
            pago = Pago()
            for name, value in request.POST.items():
                if name.startswith('cantidadCuotas'):
                    pago.cantidadCuotas=value
            pago.save()
            cantidad = pago.cantidadCuotas
            tramite.mactivo = cantidad
            tramite.save()
        return render(request, "persona/propietario/seleccionar_modo_pago.html",
                      {'tramite': tramite,'estilos': estilos})

    return render(request, 'persona/propietario/elegir_financiacion_propietario.html',{'tramite': tramite,'estilos':estilos})

def seleccionar_modo_pago(request, pk_tramite):
    estilos = ''
    usuario = request.user
    propietario = get_object_or_404(Propietario, pk=usuario.persona.propietario.pk)
    if propietario.estilo:
        estilos = propietario.estilo
    tramite = get_object_or_404(Tramite, pk=pk_tramite)
    if request.method == "POST":
        if "Guardar" in request.POST:
            pago = Pago()
            contador = 31
            fms = "%A"
            pago.cantidadCuotas = tramite.mactivo
            total = tramite.monto_a_pagar / int(pago.cantidadCuotas)
            pago.save()
            for i in range(1, int(pago.cantidadCuotas) + 1):
                cuota = Cuota(monto=total, numeroCuota=i, pago=pago)
                cuota.fechaVencimiento = date.today() + timedelta(days=contador)
                dia = cuota.fechaVencimiento.strftime(fms)
                if dia == "Sunday":
                    cuota.fechaVencimiento == date.today() + timedelta(days=contador + 1)
                else:
                    if dia == "Saturday":
                        cuota.fechaVencimiento = date.today() + timedelta(days=contador + 2)
                contador = contador + 31
                cuota.save()
                cuota.hacer("Cancelacion")
            messages.add_message(request, messages.SUCCESS, 'Financiacion registrada')
            tramite.pago = pago
            tramite.save()
        return redirect('propietario')
    return render(request, "persona/propietario/seleccionar_modo_pago.html",{'tramite': tramite, 'ctxpago': registrar_pago_propietario(request, tramite.id),
                       'estilos': estilos})

def listado_cuotas_propietario(request):
    estilos = ''
    usuario = request.user
    propietario = get_object_or_404(Propietario, pk=usuario.persona.propietario.pk)
    if propietario.estilo:
        estilos = propietario.estilo
    cuotas=Cuota.objects.en_estado(Cancelacion)
    contexto= {'cuotas':cuotas, 'estilos':estilos}
    return contexto

def elegir_tramite_propietario(request, pk_tramite):
    estilos = ''
    usuario = request.user
    propietario = get_object_or_404(Propietario, pk=usuario.persona.propietario.pk)
    if propietario.estilo:
        estilos = propietario.estilo
    tramite=get_object_or_404(Tramite,pk=pk_tramite)
    pago=tramite.pago
    cuotas=[]
    c=None
    objetos=Cuota.objects.en_estado(Cancelacion)
    for cuota in objetos:
        if cuota.fechaPago is None and cuota.pago==pago:
            c=cuota
            break;
    return render(request, 'persona/propietario/registrar_cuota_propietario.html', {'cuotas':c, 'tramite':tramite, 'estilos':estilos})

def comprobante_pago_cuota_propietario(request,pk_cuota):
    estilos = ''
    usuario = request.user
    propietario = get_object_or_404(Propietario, pk=usuario.persona.propietario.pk)
    if propietario.estilo:
        estilos = propietario.estilo
    cuota = get_object_or_404(Cuota, pk=pk_cuota)
    value = "estilo3"
    pago = cuota.pago
    tramite = get_object_or_404(Tramite, pago=pago)
    if estilos == value:
        return render(request, 'persona/propietario/comprobante_propietario_modoNocturno.html',{'cuota': cuota, 'pago':pago,'tramite':tramite,'estilos':estilos})
    return render(request, 'persona/propietario/comprobante_propietario.html',{'cuota': cuota, 'pago':pago,'tramite':tramite,'estilos':estilos})

def registrar_el_pago_tramite_propietario(request):
    estilos = ''
    usuario = request.user
    propietario = get_object_or_404(Propietario, pk=usuario.persona.propietario.pk)
    if propietario.estilo:
        estilos = propietario.estilo
    c = []
    for name, value in request.POST.items():
        if name.startswith('cuota'):
            pk = name.split('-')[1]
            c.append(pk)
    p = []
    cuotas = []
    for cs in c:
        cuota = get_object_or_404(Cuota, pk=cs)
        cuotas.append(cuota)
    pago = cuota.pago
    tramite = get_object_or_404(Tramite, pago=pago)
    contexto = {'tramite': tramite,'pago':pago, 'cuota': cuotas, 'estilos':estilos}
    return render (request,'persona/propietario/registrar_pago_tramite_propietario.html', contexto)

def listado_tramites_propietario(request):
    estilos = ''
    usuario = request.user
    propietario = get_object_or_404(Propietario, pk=usuario.persona.propietario.pk)
    if propietario.estilo:
        estilos = propietario.estilo
    tramites = Tramite.objects.filter(propietario = propietario, pago_id__isnull=False, monto_pagado__gt=0)
    contexto={'tramites':tramites, 'estilos':estilos}
    return contexto

def listado_comprobantes_propietario(request,pk_tramite):
    estilos = ''
    usuario = request.user
    propietario = get_object_or_404(Propietario, pk=usuario.persona.propietario.pk)
    if propietario.estilo:
        estilos = propietario.estilo
    tramite = get_object_or_404(Tramite, pk=pk_tramite)
    pago = tramite.pago
    canceladas = []
    cuotas = Cuota.objects.en_estado(Cancelada)
    value = "estilo3"
    for cuota in cuotas:
        if cuota.pago == pago:
            canceladas.append(cuota)
    if canceladas is None:
        messages.add_message(request, messages.WARNING, 'No hay pagos registrados para el tramite seleccionado.')
    if estilos == value:
        return render(request, 'persona/propietario/factura_parcial_propietario_modoNocturno.html', {'cuotas': canceladas, 'tramite': tramite, 'pago': pago, 'estilos': estilos})
    return render(request, 'persona/propietario/factura_parcial_propietario.html',
                  {'cuotas': canceladas, 'tramite': tramite, 'pago': pago, 'estilos': estilos})

def planilla_visado_impresa_propietario(request, pk_planilla):
    estilos = ''
    usuario = request.user
    propietario = get_object_or_404(Propietario, pk=usuario.persona.propietario.pk)
    if propietario.estilo:
        estilos = propietario.estilo
    value = "estilo3"
    filas = FilaDeVisado.objects.all()
    columnas = ColumnaDeVisado.objects.all()
    planilla = get_object_or_404(PlanillaDeVisado, id=pk_planilla)
    tramite = get_object_or_404(Tramite, pk=planilla.tramite_id)
    try:

        elementos = planilla.elementos.filter(activo=1)
        items = planilla.items.filter(activo=1)
        obs = planilla.observacion
        contexto={'tramite': tramite,
                  'planilla': planilla,
                  'filas': filas,
                  'columnas': columnas,
                  'elementos': elementos,
                  'items': items,
                  'obs': obs,
                  'estilos':estilos,
                  }
        if estilos == value:
            return render(request, 'persona/propietario/planilla_visado_impresa_propietarioModoNocturno.html',
                          contexto)

        return render(request, 'persona/propietario/planilla_visado_impresa_propietario.html',contexto)
    except:
         contexto = {
             'tramite': tramite,
    #         'planilla': planilla,
             'filas': filas,
             'columnas': columnas,
             'obs': obs,
             'estilos':estilos,
         }
         if estilos == value:
             return render(request, 'persona/propietario/planilla_visado_impresa_propietarioModoNocturno.html',
                           contexto)

         return render(request, 'persona/propietario/planilla_visado_impresa_propietario.html', contexto)

def planilla_inspeccion_impresa_propietario(request, pk_tramite):
    estilos = ''
    usuario = request.user
    propietario = get_object_or_404(Propietario, pk=usuario.persona.propietario.pk)
    if propietario.estilo:
        estilos = propietario.estilo
    value = "estilo3"
    items = ItemInspeccion.objects.all()
    categorias = CategoriaInspeccion.objects.all()
    try:
        planilla = get_object_or_404(PlanillaDeInspeccion, id=pk_tramite) # en todos los metodos el pk_tramite es el pk de la planilla no el del tramite
        tramite = get_object_or_404(Tramite, id=planilla.tramite_id)
        detalles = planilla.detalles.all()
        contexto = {
            'tramite': tramite,
            'planilla': planilla,
            'items': items,
            'categorias': categorias,
            'detalles': detalles,
            'estilos':estilos,
        }
        if estilos == value:
            return render(request, 'persona/propietario/planilla_inspeccion_impresa_propietarioModoNocturno.html',
                          contexto)

        return render(request, 'persona/propietario/planilla_inspeccion_impresa_propietario.html', contexto)

    except:
        contexto = {
       #     'tramite':tramite,
        #    'planilla': planilla,
            'items': items,
            'categorias': categorias,
            'estilos':estilos,
        }
        if estilos == value:
            return render(request, 'persona/propietario/planilla_inspeccion_impresa_propietarioModoNocturno.html',
                          contexto)

    return render(request, 'persona/propietario/planilla_inspeccion_impresa_propietario.html', contexto)


#-------------------------------------------------------------------------------------------------------------------
#profesional -------------------------------------------------------------------------------------------------------

from planilla_inspeccion.models import PlanillaDeInspeccion

@login_required(login_url="login")
@grupo_requerido('profesional')
def mostrar_profesional(request):
    usuario = request.user
    tipos_de_documentos_requeridos = TipoDocumento.get_tipos_documentos_para_momento(TipoDocumento.INICIAR)
    FormularioDocumentoSet = FormularioDocumentoSetFactory(tipos_de_documentos_requeridos)
    inicial = metodo(tipos_de_documentos_requeridos)
    documento_set = FormularioDocumentoSet(initial=inicial)
    tramite_form = FormularioIniciarTramite(initial={'profesional':usuario.persona.profesional.pk})
    propietario_form = FormularioPropietario()
    propietario = None
    prop=0
    propietarioValido=0
    documentosValidos=0
    crearPropietario=0
    dniIncorrecto=0
    if request.method == "POST":
        personas = Persona.objects.filter(dni=request.POST["propietario"])
        persona = personas.exists() and personas.first() or None
        propietario_form = FormularioPropietario(request.POST)
        documento_set = FormularioDocumentoSet(request.POST, request.FILES)

        if documento_set.is_valid():
            documentosValidos = 1
        try:
                propietario = persona.propietario
                crearPropietario=0
                propietarioValido=1
        except:
                crearPropietario=1

        if documentosValidos == 1 and crearPropietario == 1:
            try:
                nuevoPropietario = request.POST["dni"]
                dniPropietario = request.POST['propietario']
                if dniPropietario == nuevoPropietario:
                    propietario = propietario_form.crear(persona)
                    propietarioValido = 1
                    dniIncorrecto=0
                else:
                    propietarioValido = 0
                    dniIncorrecto=1
            except:
                propietarioValido = 0
        tramite_form = FormularioIniciarTramite(request.POST)
        if propietario is not None and tramite_form.is_valid() and propietarioValido == 1 and documentosValidos == 1 :
            tramite = tramite_form.save(propietario=propietario, commit=False)
            lista=[]
            for docForm in documento_set:
               lista.append(docForm.save(commit=False))
            Tramite.new(
                usuario,
                propietario,
                usuario.persona.profesional,
                request.POST['tipo_obra'],
                request.POST['medidas'],
                request.POST['domicilio'],
                request.POST['parcela'],
                request.POST['circunscripcion'],
                request.POST['manzana'],
                request.POST['sector'],
                lista
            )
            messages.add_message(request, messages.SUCCESS,
                                 'Solicitud de inicio de tramite exitosa.')
            tramite_form = FormularioIniciarTramite(initial={'profesional':usuario.persona.profesional.pk})
            propietario_form = None
            return redirect('profesional')
        else:
            if propietarioValido == 0  and dniIncorrecto == 1:
                messages.add_message(request, messages.WARNING, 'Los numeros de DNI ingresados no coinciden')
                prop=1
            elif documentosValidos == 0 and propietario is None :
                messages.add_message(request, messages.WARNING, 'El propietario ingresado no existe, debe darlo de alta para iniciar al tramite. ')
                messages.add_message(request, messages.WARNING, 'Debe ingresar los documentos faltantes')
                prop=1
            elif  propietario is None:
                messages.add_message(request, messages.WARNING, 'El propietario ingresado no existe, debe darlo de alta para iniciar al tramite.')
                prop=1
            elif documentosValidos == 0:
                messages.add_message(request, messages.WARNING, 'Debe ingresar los documentos faltantes')
            else:
                prop=0
    else:
        propietario_form = None
    contexto = {
        'documentos_requeridos': tipos_de_documentos_requeridos,
        'ctxtramitesprofesional': listado_tramites_de_profesional(request),
        'tramite_form': tramite_form,
        'propietario_form': propietario_form,
        'documento_set': documento_set,
        'ctxtramcorregidos':tramites_corregidos(request),
        'prop': prop,
    }
    return render(request, 'persona/profesional/profesional.html', contexto)

def listado_tramites_de_profesional(request):
    tramites = Tramite.objects.all()
    usuario = request.user
    persona = Persona.objects.get(usuario__isnull=False, usuario_id=usuario)
    tramites_de_profesional = filter(lambda tramite: (tramite.profesional == persona.profesional), tramites)
    contexto = {'tramites_de_profesional': tramites_de_profesional}
    return contexto

def tramites_corregidos(request):
    tram_corregidos=[]
    usuario = request.user
    persona = Persona.objects.get(usuario__isnull=False, usuario_id=usuario)
    tramites = Tramite.objects.filter(profesional_id=persona.profesional)
    tipo = 4
    for t in tramites:
        try:
            if t.estado().tipo==tipo:
                tram_corregidos.append(t)
        except:
            pass
    contexto = {'tramites': tram_corregidos}
    return contexto

def tramites_corregidos_administrativo(request):
    anterior=4
    tram_corregidos=[]
    tram=Tramite.objects.en_estado(Iniciado)
    for t in tram:
        estados=t.estadosTramite()
        try:
            e=estados.last()
            estadoPrevio = e.previo()
            estadoSiguiente = e.siguiente()
            if (estadoSiguiente is None  and estadoPrevio.tipo == anterior):
                tram_corregidos.append(t)
        except:
            pass
    contexto = {'tramites': tram_corregidos}
    return contexto

def ver_documentos_tramite_profesional(request, pk_tramite):
    tramite = get_object_or_404(Tramite, pk=pk_tramite)
    contexto0 = {'tramite': tramite}
    estados_de_tramite = tramite.estadosTramite()
    contexto1 = {'estados_del_tramite': estados_de_tramite}
    fechas_del_estado = [];
    for est in estados_de_tramite:
        fechas_del_estado.append(est.timestamp.strftime("%d/%m/%Y"));
    return render(request, 'persona/profesional/vista_de_documentos.html', {"tramite": contexto0, "estadosp": contexto1, "fecha":fechas_del_estado})

def profesional_solicita_final_obra(request, pk_tramite):
    tramite = get_object_or_404(Tramite, pk=pk_tramite)
    try:
        if tramite.estados.filter(tipo=8):
            messages.add_message(request, messages.WARNING, "La solicitud de final de obra ya fue otorgada anteriormente")
        else:
            tramite.hacer(Tramite.SOLICITAR_FINAL_OBRA, request.user)
            messages.add_message(request, messages.SUCCESS, 'final de obra solicitado.')
    except:
        messages.add_message(request, messages.ERROR, 'No puede solicitar el final de obra para ese tramite.')
    finally:
        return redirect('profesional')

def ver_documentos_corregidos(request, pk_tramite):
    if request.method == "POST":
       if  "Enviar correccion de administrativo" in request.POST:
            tipo= "correccion visado" #deberia venir por post en un input hidden
            documentos = Documento.objects.filter(tramite_id=pk_tramite)
            enviar_correccioness(request, pk_tramite,tipo)
    else:
        tramite = get_object_or_404(Tramite, pk=pk_tramite)
        planillas = PlanillaDeVisado.objects.filter(
            tramite_id=tramite.id)  # busca las planillas que tengan el id del tramite
        if (len(planillas) > 1):
            aux = planillas[0]
            for p in planillas:
                if (p.id > aux.id):  # obtiene el ultimo visado del tramite
                    plan = p
                else:
                    plan = aux
            planilla = get_object_or_404(PlanillaDeVisado,
                                         id=plan.id)  # PlanillaDeVisado.objects.filter(tramite_id=tramite.id)# busca las planillas que tengan el id del tramite
        else:
            try:
                planilla = PlanillaDeVisado.objects.get(
                    tramite_id=pk_tramite)  # PlanillaDeVisado.objects.filter(tramite_id=tramite.id)# busca las planillas que tengan el id del tramite
            except:
                planilla = "No hay planilla para mostrar"
        filas = FilaDeVisado.objects.all()
        columnas = ColumnaDeVisado.objects.all()
        try:
            obs = planilla.observacion
            elementos = planilla.elementos.all()
            items = planilla.items.all()
            contexto = {
                'tramite': tramite,
                'planilla': planilla,
                'filas': filas,
                'columnas': columnas,
                'items': items,
                'elementos': elementos,
                'obs': obs,
            }
            return render(request, 'persona/profesional/ver_documentos_corregidos.html', contexto)
        except:
            contexto = {
                'tramite': tramite,
                'mensaje': planilla,
            }
            return render(request, 'persona/profesional/ver_documentos_corregidos.html', contexto)

    return redirect('profesional')


def enviar_correccioness(request, pk_tramite, tipo):
    usuario = request.user
    observacion = "Este tramite ya tiene los archivos corregidos cargados"
    try:
        tramite = get_object_or_404(Tramite, pk=pk_tramite)
        documento = Documento(file=request.FILES['documento'])
        tipo_documento = TipoDocumento.objects.get(nombre=tipo)
        documento.tipo_documento = tipo_documento
        documento.tramite = tramite
        documento.save()
        tramite.hacer(tramite.CORREGIR, request.user, observacion)
        messages.add_message(request, messages.SUCCESS, 'Tramite con documentos corregidos y enviados')
    except:
        messages.add_message(request, messages.ERROR, 'No se pudo enviar la correccion del tramite')
    return redirect('profesional')

def enviar_correcciones(request, pk_tramite):
    usuario = request.user
    #archivos = request.GET['msg']
    observacion = "Este tramite ya tiene los archivos corregidos cargados"
    tramite = get_object_or_404(Tramite, pk=pk_tramite)
    documento = Documento(file=request.FILES['documento'])
    documento.tramite = tramite
    documento.save()
    tramite.hacer(tramite.CORREGIR, request.user, observacion)
    messages.add_message(request, messages.SUCCESS, 'Tramite con documentos corregidos y enviados')
    return redirect('profesional')
'''
def documento_de_estado(request, pk_estado):
    estado = get_object_or_404(Estado, pk=pk_estado)
    fecha = estado.timestamp
    fecha_str = date.strftime(fecha, '%d/%m/%Y %H:%M')
    documentos = estado.tramite.documentos.all()
    documentos_fecha = filter(lambda e: (date.strftime(e.fecha, '%d/%m/%Y %H:%M') <= fecha_str), documentos)
    contexto = {'documentos_de_fecha': documentos_fecha}
    planillas = []
    inspecciones = []
    documento = estado.tramite.documentacion_para_estado(estado)
    if (estado.tipo == 1 or estado.tipo == 2):
        contexto = {'documentos': documentos_fecha}
    if (estado.tipo > 2 and estado.tipo < 5):
        for p in PlanillaDeVisado.objects.all():
            if (p.tramite.pk == estado.tramite.pk):
                planillas.append(p)
        filas = FilaDeVisado.objects.all()
        columnas = ColumnaDeVisado.objects.all()
        contexto = {
            'planillas': planillas,
            'filas': filas,
            'columnas': columnas,
        }
    if (estado.tipo > 5 and estado.tipo <= 9):
        for p in PlanillaDeInspeccion.objects.all():
            if (p.tramite.pk == estado.tramite.pk):
                inspecciones.append(p)
        items = ItemInspeccion.objects.all()
        categorias = CategoriaInspeccion.objects.all()
        contexto = {
            'inspecciones': inspecciones,
            'items': items,
            'categorias': categorias,
        }
    return render(request, 'persona/profesional/documento_de_estado.html', contexto)
'''
def documento_de_estado(request, pk_estado):
    estado = get_object_or_404(Estado, pk=pk_estado)
    documentos = estado.tramite.documentacion_para_estado(estado)
    return render(request, 'persona/profesional/documento_de_estado.html', documentos)

def planilla_visado_impresa(request, pk_tramite):
    planilla = get_object_or_404(PlanillaDeVisado,id=pk_tramite)
    tramite = get_object_or_404(Tramite, pk=planilla.tramite_id)
    filas = FilaDeVisado.objects.all()
    columnas = ColumnaDeVisado.objects.all()
    try:
        elementos = planilla.elementos.filter(activo=1)
        items = planilla.items.filter(activo=1)
        obs = planilla.observacion
        contexto={'tramite': tramite,
                  'planilla': planilla,
                  'filas': filas,
                  'columnas': columnas,
                  'elementos': elementos,
                  'items': items,
                  'obs': obs,
                  }
        return render(request, 'persona/profesional/planilla_visado_impresa.html',contexto)
    except:
         contexto = {
             'tramite': tramite,
             'planilla': planilla,
             'filas': filas,
             'columnas': columnas,
             'obs': obs,
         }
         return render(request, 'persona/profesional/planilla_visado_impresa.html', contexto)

def planilla_inspeccion_impresa(request, pk_tramite):
    planilla = get_object_or_404(PlanillaDeInspeccion, id=pk_tramite)
    tramite=get_object_or_404(Tramite, id=planilla.tramite_id)
    items = ItemInspeccion.objects.all()
    categorias = CategoriaInspeccion.objects.all()
    try:
        detalles = planilla.detalles.all()
        contexto = {
            'tramite': tramite,
            'planilla': planilla,
            'items': items,
            'categorias': categorias,
            'detalles': detalles,
        }
        return render(request, 'persona/profesional/planilla_inspeccion_impresa.html', contexto)

    except:
        contexto = {
            'tramite':tramite,
            'planilla': planilla,
            'items': items,
            'categorias': categorias,
        }
        return render(request, 'persona/profesional/planilla_inspeccion_impresa.html', contexto)


class ReporteTramitesProfesionalPdf(View):

    def get(self, request, *args, **kwargs):
        filename = "Informe de tramites.pdf"
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="%s"' % filename
        doc = SimpleDocTemplate(
            response,
            pagesize=letter,
            rightMargin=72,
            leftMargin=72,
            topMargin=15,
            bottomMargin=28,
        )
        Story = []
        styles = getSampleStyleSheet()
        styles.add(ParagraphStyle(name='Usuario', alignment=TA_RIGHT, fontName='Helvetica', fontSize=8))
        styles.add(ParagraphStyle(name='Titulo', alignment=TA_RIGHT, fontName='Helvetica', fontSize=18))
        styles.add(ParagraphStyle(name='Subtitulo', alignment=TA_RIGHT, fontName='Helvetica', fontSize=12))
        usuario = 'Usuario: ' + request.user.username + ' -  Fecha:' + str(datetime.date.today())
        Story.append(Paragraph(usuario, styles["Usuario"]))
        Story.append(Spacer(0, cm * 0.15))
        im0 = Image(settings.MEDIA_ROOT + '/imagenes/espacioPDF.png', width=490, height=3)
        im0.hAlign = 'CENTER'
        Story.append(im0)
        titulo = 'SISTEMA OBRAS PARTICULARES'
        Story.append(Paragraph(titulo, styles["Titulo"]))
        Story.append(Spacer(0, cm * 0.20))
        subtitulo = 'Reporte de tramites'
        Story.append(Paragraph(subtitulo, styles["Subtitulo"]))
        Story.append(Spacer(0, cm * 0.15))
        Story.append(im0)
        Story.append(Spacer(0, cm * 0.5))
        encabezados = ('PLANILLA DE VISADO')
        detalles = [(columna, fila, '*')
           for fila in
             FilaDeVisado.objects.all()
                    for columna in ColumnaDeVisado.objects.all()
                        for item in ItemDeVisado.objects.filter(activo=True)
                            if fila == item.fila_de_visado and columna == item.columna_de_visado]
        detalle_orden = Table([encabezados] + detalles, colWidths=[5 * cm, 7 * cm, 6 * cm, 6 * cm, 4 * cm, 4 * cm])
        detalle_orden.setStyle(TableStyle(
            [
                ('ALIGN', (0, 0), (0, 0), 'CENTER'),
                ('GRID', (0, 0), (-1, -1), 1, colors.gray),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('LINEBELOW', (0, 0), (-1, 0), 2, colors.darkblue),
                ('BACKGROUND', (0, 0), (-1, 0), colors.dodgerblue),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ]
        ))
        detalle_orden.hAlign = 'CENTER'
        Story.append(detalle_orden)
        doc.build(Story)
        return response

#-------------------------------------------------------------------------------------------------------------------
#administrativo ----------------------------------------------------------------------------------------------------

@login_required(login_url="login")
@grupo_requerido('administrativo')
def mostrar_administrativo(request):

    contexto = {
         "ctxprofesional": profesional_list(request),
         "ctxpropietario": propietario_list(request),
         "ctxtramitesiniciados": listado_de_tramites_iniciados(request),
         "ctxtramitescorregidos": tramite_corregidos_list(request),
         "ctxsolicitudesfinalobra": solicitud_final_obra_list(request),
       # "ctxpago": registrar_pago_tramite(request),
         "ctxlistprofesional": listado_profesionales(request),
         'ctxtramcorregidosadministrativo': tramites_corregidos_administrativo(request)

    }
    return render(request, 'persona/administrativo/administrativo.html', contexto)

def profesional_list(request):
    profesionales = Persona.objects.filter(usuario__isnull=True,  profesional_id__isnull = False)
#    profesionales = filter(lambda persona: (persona.usuario is None and persona.profesional is not None), personas)
    contexto = {'personas': profesionales}
    return contexto

def propietario_list(request):
    propietarios_sin_usuario = Propietario.objects.filter(persona__usuario__isnull=True, persona__isnull = False)
    #propietarios_sin_usuario = filter(lambda propietario: (propietario.persona.usuario is None and propietario.persona is not None ), propietarios)
    contexto = {'propietarios': propietarios_sin_usuario}
    return contexto
#from time import time
def listado_de_tramites_iniciados(request):
    tramites = Tramite.objects.en_estado(Iniciado)
    return {'tramites': tramites}

def tramite_corregidos_list(request):
    tramites = Tramite.objects.en_estado(Corregido)
    return {'tramites': tramites}

def solicitud_final_obra_list(request):
    estadosFinalObraSolicitado=Estado.objects.filter(tipo=8).select_related()
    tramitesHabilitados=[]
    for t in estadosFinalObraSolicitado:
        estados=t.tramite.estadosTramite().values_list('tipo', flat=True)
        if  9 not in estados and 7 in estados :
            tramitesHabilitados.append(t.tramite)

    # for e in estadosFinalObraSolicitado:
    #     estado = e.tramite.estados.all()
    #     encontrado=[True for i in estado if i.tipo==9]
    #     inspeccionado=[True for i in estado if i.tipo==7]
    #     if not encontrado and inspeccionado:
    #         tramites_id.append(e.tramite_id)
    # for t in tramites:
    #     for t_id in tramites_id:
    #         if t_id ==t.id:
    #             tramitesHabilitados.append(t)
    contexto = {'tramites': tramitesHabilitados}
    return contexto

def registrar_pago_tramite(request):
    if request.method == "POST":
        archivo_pago_form = FormularioArchivoPago(request.POST, request.FILES)
        if archivo_pago_form.is_valid():
            Pago.procesar_pagos(request.FILES['pagos'])
    else:
        archivo_pago_form = FormularioArchivoPago()
    return archivo_pago_form

def crear_usuario(request, pk_persona):
    usuario = request.user
    persona = get_object_or_404(Persona, pk=pk_persona)
    creado, password, usuario_creado = persona.crear_usuario()
    if creado:
        messages.add_message(request, messages.SUCCESS, 'usuario creado.')
        # Mandar correo al  nuevo usuario con su usurio y clave
        print("Mando correo de creado")
        print(password)
        send_mail(
            'Usuario habilitado',
            'Usted ya puede acceder al sistema: Nombre de usuario: '+persona.mail+' password: '+password,
            'infosopunpsjb@gmail.com',
            [persona.mail],
            fail_silently=False,
        )
    else:
        print("Mando correo informando que se cambio algo en su cuenta de usuario")
    return redirect(usuario.get_view_name())

def habilitar_final_obra(request, pk_tramite):
    tramite = get_object_or_404(Tramite, pk=pk_tramite)
    try:
        tramite.hacer(tramite.FINALIZAR, request.user)
        messages.add_message(request, messages.SUCCESS, 'final de obra habilitado.')
    except:
        messages.add_message(request, messages.ERROR, 'No puede otorgar final de obra total para ese tramite.')
    finally:
        return redirect('administrativo')
def aceptar_tramite(request, pk_tramite):
    tramite = get_object_or_404(Tramite, pk=pk_tramite)
    tramite.hacer(tramite.ACEPTAR, request.user)
    messages.add_message(request, messages.SUCCESS, "Tramite aceptado")
    return redirect('administrativo')

def rechazar_tramite(request, pk_tramite):
    tramite = get_object_or_404(Tramite, pk=pk_tramite)
    tramite.hacer(tramite.RECHAZAR, request.user, request.GET["msg"])
    messages.add_message(request, messages.WARNING, 'Tramite rechazado.')
    return redirect('administrativo')

class ver_un_certificado(DetailView):
    model = Persona
    template_name = 'persona/administrativo/ver_certificado_profesional.html'
    def dispatch(self, *args, **kwargs):
        return super(ver_un_certificado, self).dispatch(*args, **kwargs)

def ver_documentos_tramite_administrativo(request, pk_tramite):
    tramite = get_object_or_404(Tramite, pk=pk_tramite)
    return render(request, 'persona/administrativo/vista_de_documentos_administrativo.html', {'tramite': tramite})

def documentos_administrativo(request, pk_tramite):
    tramite = get_object_or_404(Tramite, pk=pk_tramite)
    contexto0 = {'tramite': tramite}
    estados_de_tramite = tramite.estadosTramite()
    contexto1 = {'estados_del_tramite': estados_de_tramite}
    fechas_del_estado = [];
    for est in estados_de_tramite:
        fechas_del_estado.append(est.timestamp.strftime("%d/%m/%Y"));
    return render(request, 'persona/administrativo/documentos_administrativo.html', {"tramite": contexto0, "estadosp": contexto1, "fecha":fechas_del_estado})

def documento_de_estado_administrativo(request, pk_estado):
    estado = get_object_or_404(Estado, pk=pk_estado)
    documentos = estado.tramite.documentacion_para_estado(estado)
    return render(request, 'persona/administrativo/documento_de_estado_administrativo.html', documentos)

'''
def documento_de_estado_administrativo(request, pk_estado):
    estado = get_object_or_404(Estado, pk=pk_estado)
    fecha = estado.timestamp
    fecha_str = date.strftime(fecha, '%d/%m/%Y %H:%M')
    documentos = estado.tramite.documentos.all()
    documentos_fecha = filter(lambda e: (date.strftime(e.fecha, '%d/%m/%Y %H:%M') == fecha_str), documentos)
    contexto = {'documentos_de_fecha': documentos_fecha}
    planillas = []
    inspecciones = []
    documento = estado.tramite.documentacion_para_estado(estado)
    if (estado.tipo == 1 or estado.tipo == 2):
        contexto = {'documentos': documentos}
    if (estado.tipo > 2 and estado.tipo < 5):
        for p in PlanillaDeVisado.objects.all():
            if (p.tramite.pk == estado.tramite.pk):
                planillas.append(p)
        filas = FilaDeVisado.objects.all()
        columnas = ColumnaDeVisado.objects.all()
        contexto = {
            'planillas': planillas,
            'filas': filas,
            'columnas': columnas,
        }
    if (estado.tipo >= 5 and estado.tipo < 8):
        for p in PlanillaDeInspeccion.objects.all():
            if (p.tramite.pk == estado.tramite.pk):
                inspecciones.append(p)
        items = ItemInspeccion.objects.all()
        categorias = CategoriaInspeccion.objects.all()
        contexto = {
            'inspecciones': inspecciones,
            'items': items,
            'categorias': categorias,
        }
    return render(request, 'persona/administrativo/documento_de_estado_administrativo.html', contexto)
'''
def listado_profesionales(request):
    tramites = Tramite.objects.all().values_list('profesional_id',flat=True) # puse con inspeccion solo para fines de mostrar algo
    personas = Profesional.objects.filter(id__in=tramites)
    contexto = {
        "profesionales": personas}
    return contexto

def lista_profesionales_imprimible(request):
    personas = Profesional.objects.all()
    profesionales = filter(lambda persona: (persona is not None), personas)
    contexto = {'profesionales': personas}
    return render(request, 'persona/administrativo/lista_profesionales_imprimible.html', contexto)

class ReporteProfesionalExcel(TemplateView):
    def get(self, request, *args, **kwargs):
        profesionales = Profesional.objects.all()
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'REPORTE DE PROFESIONALES'
        ws.merge_cells('B1:G1')
        ws['B2'] = 'NOMBRE'
        ws['C2'] = 'APELLIDO'
        ws['D2'] = 'DNI'
        ws['E2'] = 'TELEFONO'
        ws['F2'] = 'PROFESION'
        ws['G2'] = 'CATEGORIA'
        ws['H2'] = 'MATRICULA'
        ws['I2'] = 'DOMICILIO'
        ws['J2'] = 'MAIL'
        cont = 3
        for profesional in profesionales:
            ws.cell(row=cont, column=2).value = str(profesional.persona.nombre)
            ws.cell(row=cont, column=3).value = str(profesional.persona.apellido)
            ws.cell(row=cont, column=4).value = str(profesional.persona.dni)
            ws.cell(row=cont, column=5).value = str(profesional.persona.telefono)
            ws.cell(row=cont, column=6).value = str(profesional.profesion)
            ws.cell(row=cont, column=7).value = profesional.categoria
            ws.cell(row=cont, column=8).value = profesional.matricula
            ws.cell(row=cont, column=9).value = str(profesional.persona.domicilio)
            ws.cell(row=cont, column=10).value = str(profesional.persona.mail)
            cont = cont + 1
        nombre_archivo = "ReportePersonasExcel.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response

class ReporteProfesionalesPdf(View):

    def get(self, request, *args, **kwargs):

        filename = "Informe de profesionales.pdf"
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="%s"' % filename
        doc = SimpleDocTemplate(
            response,
            pagesize=letter,
            rightMargin=72,
            leftMargin=72,
            topMargin=15,
            bottomMargin=28,
        )
        Story = []
        styles = getSampleStyleSheet()
        styles.add(ParagraphStyle(name='Usuario', alignment=TA_RIGHT, fontName='Helvetica', fontSize=8))
        styles.add(ParagraphStyle(name='Titulo', alignment=TA_RIGHT, fontName='Helvetica', fontSize=18))
        styles.add(ParagraphStyle(name='Subtitulo', alignment=TA_RIGHT, fontName='Helvetica', fontSize=12))
        fecha = timezone.now().strftime('%Y-%m-%d')
        usuario = 'Usuario: ' + request.user.username + ' -  Fecha:' + str(fecha)
        Story.append(Paragraph(usuario, styles["Usuario"]))
        Story.append(Spacer(0, cm * 0.1))
        im0 = Image(settings.MEDIA_ROOT + '/imagenes/espacioPDF.png', width=490, height=3)
        im0.hAlign = 'CENTER'
        Story.append(im0)
        titulo = 'SISTEMA OBRAS PARTICULARES'
        Story.append(Paragraph(titulo, styles["Titulo"]))
        Story.append(Spacer(0, cm * 0.20))
        subtitulo = 'Listado de Profesionales activos'
        Story.append(Paragraph(subtitulo, styles["Subtitulo"]))
        Story.append(Spacer(0, cm * 0.15))
        Story.append(im0)
        Story.append(Spacer(0, cm * 0.5))
        encabezados = ('Nombre', 'Apellido', 'Telefono','Profesion',
                       'Matricula','Domicilio','correo')
        detalles = [(profesional.persona.nombre, profesional.persona.apellido,
                     profesional.persona.dni,
                     profesional.persona.telefono, profesional.profesion,
                     profesional.matricula,
                     profesional.persona.domicilio,
                     profesional.persona.mail) for
                    profesional in
                    Profesional.objects.all()
                    ]
        detalle_orden = Table([encabezados] + detalles, colWidths=[2 * cm, 2 * cm, 2 * cm, 2 * cm, 2 * cm,
                                                                   4 * cm, 6 * cm])
        detalle_orden.setStyle(TableStyle(
            [
                ('ALIGN', (0, 0), (0, 0), 'CENTER'),
                ('GRID', (0, 0), (-1, -1), 1, colors.gray),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('LINEBELOW', (0, 0), (-1, 0), 2, colors.darkblue),
                ('BACKGROUND', (0, 0), (-1, 0), colors.dodgerblue),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ]
        ))
        detalle_orden.hAlign = 'CENTER'
        Story.append(detalle_orden)
        doc.build(Story)
        return response

#reporte tramites iniciados

class ReporteTramitesIniciadosExcel(TemplateView):
    def get(self, request, *args, **kwargs):
        tramites = Tramite.objects.en_estado(Iniciado)
        wb = Workbook()
        ws = wb.active
        ws['B1'] = 'REPORTE DE TRAMITES INICIADOS'
        ws.merge_cells('B1:G1')
        ws.merge_cells('B2:C2')
        ws.merge_cells('D2:E2')
        ws.merge_cells('F2:G2')
        ws.merge_cells('H2:I2')
        ws.merge_cells('J2:K2')
        ws['B2'] = 'NRO DE TRAMITE'
        ws['D2'] = 'PROPIETARIO'
        ws['F2'] = 'PROFESIONAL'
        ws['H2'] = 'MEDIDAS'
        ws['J2'] = 'TIPO'
        cont = 3
        for tramite in tramites:
            ws.cell(row=cont, column=2).value = str(tramite.id)
            ws.cell(row=cont, column=4).value = str(tramite.propietario)
            ws.cell(row=cont, column=6).value = str(tramite.profesional)
            ws.cell(row=cont, column=8).value = str(tramite.medidas)
            ws.cell(row=cont, column=10).value = str(tramite.tipo_obra)
            cont = cont + 1
        nombre_archivo = "ReportePersonasExcel.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response

class ReporteTramitesIniciadosPdf(View):
    def get(self, request, *args, **kwargs):
        filename = "Informe de tramites iniciados " + datetime.datetime.now().strftime("%d/%m/%Y") + " .pdf"
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="%s"' % filename
        doc = SimpleDocTemplate(
            response,
            pagesize=letter,
            rightMargin=72,
            leftMargin=72,
            topMargin=15,
            bottomMargin=28,
        )
        Story = []
        styles = getSampleStyleSheet()
        styles.add(ParagraphStyle(name='Usuario', alignment=TA_RIGHT, fontName='Helvetica', fontSize=8))
        styles.add(ParagraphStyle(name='Titulo', alignment=TA_RIGHT, fontName='Helvetica', fontSize=18))
        styles.add(ParagraphStyle(name='Subtitulo', alignment=TA_RIGHT, fontName='Helvetica', fontSize=12))
        usuario = 'Usuario: ' + request.user.username + ' - Fecha: ' + datetime.datetime.now().strftime("%Y/%m/%d")
        Story.append(Paragraph(usuario, styles["Usuario"]))
        Story.append(Spacer(0, cm * 0.15))
        im0 = Image(settings.MEDIA_ROOT + '/imagenes/espacioPDF.png', width=490, height=3)
        im0.hAlign = 'CENTER'
        Story.append(im0)
        titulo = 'SISTEMA OBRAS PARTICULARES'
        Story.append(Paragraph(titulo, styles["Titulo"]))
        Story.append(Spacer(0, cm * 0.20))
        subtitulo = 'Reporte De Tramites Iniciados'
        Story.append(Paragraph(subtitulo, styles["Subtitulo"]))
        Story.append(Spacer(0, cm * 0.15))
        Story.append(im0)
        Story.append(Spacer(0, cm * 0.5))
        encabezados = ('NRO', 'PROPIETARIO', 'PROFESIONAL', 'MEDIDAS', 'TIPO')
        detalles = [(tramite.id, tramite.propietario, tramite.profesional, tramite.medidas,
                        tramite.tipo_obra)
                    for
                    tramite in
                    Tramite.objects.en_estado(Iniciado)]
        detalle_orden = Table([encabezados] + detalles, colWidths=[2 * cm, 3 * cm, 3 * cm, 3 * cm])
        detalle_orden.setStyle(TableStyle(
            [
                ('ALIGN', (0, 0), (0, 0), 'CENTER'),
                ('GRID', (0, 0), (-1, -1), 1, colors.gray),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('LINEBELOW', (0, 0), (-1, 0), 2, colors.darkblue),
                ('BACKGROUND', (0, 0), (-1, 0), colors.dodgerblue),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ]
        ))
        detalle_orden.hAlign = 'CENTER'
        Story.append(detalle_orden)
        doc.build(Story)
        return response

#reporte tramites corregidos

class ReporteTramitesCorregidosExcel(TemplateView):
    def get(self, request, *args, **kwargs):
        tram = Tramite.objects.en_estado(Iniciado)
        anterior = 4
        tram_corregidos = []
        for t in tram:
            estados = t.estadosTramite()
            try:
                e = estados.last()
                estadoPrevio = e.previo()
                estadoSiguiente = e.siguiente()
                if (estadoSiguiente is None and estadoPrevio.tipo == anterior):
                    tram_corregidos.append(t)
            except:
                pass
        wb = Workbook()
        ws = wb.active
        ws['B1'] = 'REPORTE DE TRAMITES CORREGIDOS'
        ws.merge_cells('B1:F1')
        ws.merge_cells('C1:D1')
        ws.merge_cells('E1:F1')
        ws.merge_cells('H1:I1')
        ws['B2'] = 'NRO'
        ws['C2'] = 'PROPIETARIO'
        ws['E2'] = 'PROFESIONAL'
        ws['G2'] = 'MEDIDAS'
        ws['H2'] = 'TIPO'
        cont = 3
        for tramite in tram_corregidos:
            ws.cell(row=cont, column=2).value = str(tramite.id)
            ws.cell(row=cont, column=3).value = str(tramite.propietario)
            ws.cell(row=cont, column=5).value = str(tramite.profesional)
            ws.cell(row=cont, column=7).value = str(tramite.medidas)
            ws.cell(row=cont, column=8).value = str(tramite.tipo_obra)
            cont = cont + 1
        nombre_archivo = "ReportePersonasExcel.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response

class ReporteTramitesCorregidosPdf(View):
    def get(self, request, *args, **kwargs):
        filename = "Informe de tramites corregidos " + datetime.datetime.now().strftime("%d/%m/%Y") + " .pdf"
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="%s"' % filename
        doc = SimpleDocTemplate(
            response,
            pagesize=letter,
            rightMargin=72,
            leftMargin=72,
            topMargin=15,
            bottomMargin=28,
        )
        Story = []
        styles = getSampleStyleSheet()
        styles.add(ParagraphStyle(name='Usuario', alignment=TA_RIGHT, fontName='Helvetica', fontSize=8))
        styles.add(ParagraphStyle(name='Titulo', alignment=TA_RIGHT, fontName='Helvetica', fontSize=18))
        styles.add(ParagraphStyle(name='Subtitulo', alignment=TA_RIGHT, fontName='Helvetica', fontSize=12))
        usuario = 'Usuario: ' + request.user.username + ' - Fecha: ' + datetime.datetime.now().strftime("%Y/%m/%d")
        Story.append(Paragraph(usuario, styles["Usuario"]))
        Story.append(Spacer(0, cm * 0.15))
        im0 = Image(settings.MEDIA_ROOT + '/imagenes/espacioPDF.png', width=490, height=3)
        im0.hAlign = 'CENTER'
        Story.append(im0)
        titulo = 'SISTEMA OBRAS PARTICULARES'
        Story.append(Paragraph(titulo, styles["Titulo"]))
        Story.append(Spacer(0, cm * 0.20))
        subtitulo = 'Reporte De Tramites Corregidos'
        Story.append(Paragraph(subtitulo, styles["Subtitulo"]))
        Story.append(Spacer(0, cm * 0.15))
        Story.append(im0)
        Story.append(Spacer(0, cm * 0.5))
        encabezados = ('NRO', 'PROPIETARIO', 'PROFESIONAL', 'MEDIDAS', 'TIPO')
        tram = Tramite.objects.en_estado(Iniciado)
        anterior = 4
        tram_corregidos = []
        for t in tram:
            estados = t.estadosTramite()
            try:
                e = estados.last()
                estadoPrevio = e.previo()
                estadoSiguiente = e.siguiente()
                if (estadoSiguiente is None and estadoPrevio.tipo == anterior):
                    tram_corregidos.append(t)
            except:
                pass
        detalles = [(tramite.id, tramite.propietario, tramite.profesional, tramite.medidas,
                        tramite.tipo_obra)
                    for
                    tramite in
                    tram_corregidos]
        detalle_orden = Table([encabezados] + detalles, colWidths=[2 * cm, 3 * cm, 3 * cm, 3 * cm])
        detalle_orden.setStyle(TableStyle(
            [
                ('ALIGN', (0, 0), (0, 0), 'CENTER'),
                ('GRID', (0, 0), (-1, -1), 1, colors.gray),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('LINEBELOW', (0, 0), (-1, 0), 2, colors.darkblue),
                ('BACKGROUND', (0, 0), (-1, 0), colors.dodgerblue),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ]
        ))
        detalle_orden.hAlign = 'CENTER'
        Story.append(detalle_orden)
        doc.build(Story)
        return response

    #reporte listado profesionales activos

class ReporteProfesionalesActivosExcel(TemplateView):
    def get(self, request, *args, **kwargs):
        profesionales = Tramite.objects.select_related().values( 'profesional__persona__nombre',
                                                                'profesional__persona__apellido',
                                                                'profesional__persona__telefono',
                                                                'profesional__profesion',
                                                                'profesional__matricula',
                                                                'profesional__persona__domicilio',
                                                                'profesional__persona__mail').all().distinct()  # puse con inspeccion solo para fines de mostrar algo
        wb = Workbook()
        ws = wb.active
        ws['B1'] = 'REPORTE DE PROFESIONALES ACTIVOS'
        ws.merge_cells('B1:G1')
        ws.merge_cells('B2:C2')
        ws.merge_cells('D2:E2')
        ws.merge_cells('F2:G2')
        ws.merge_cells('H2:I2')
        ws.merge_cells('J2:K2')
        ws.merge_cells('L2:N2')
        ws.merge_cells('O2:P2')
        ws['B2'] = 'NOMBRE'
        ws['D2'] = 'APELLIDO'
        ws['F2'] = 'TELEFONO'
        ws['H2'] = 'MATRICULA'
        ws['J2'] = 'DOMICILIO'
        ws['L2'] = 'PROFESION'
        ws['O2'] = 'MAIL'
        cont = 3
        for profesional in profesionales:
            ws.cell(row=cont, column=2).value = profesional['profesional__persona__nombre']
            ws.cell(row=cont, column=4).value = profesional['profesional__persona__apellido']
            ws.cell(row=cont, column=6).value = profesional['profesional__persona__telefono']
            ws.cell(row=cont, column=8).value = profesional['profesional__matricula']
            ws.cell(row=cont, column=10).value = profesional['profesional__persona__domicilio']
            ws.cell(row=cont, column=12).value = profesional['profesional__profesion']
            ws.cell(row=cont, column=15).value = profesional['profesional__persona__mail']
            cont = cont + 1
        nombre_archivo = "ReportePersonasExcel.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response

class ReporteProfesionalesActivosPdf(View):
    def get(self, request, *args, **kwargs):

        filename = "Informe de profesionales activos.pdf"
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="%s"' % filename
        doc = SimpleDocTemplate(
            response,
            pagesize=letter,
            rightMargin=72,
            leftMargin=72,
            topMargin=15,
            bottomMargin=28,
        )
        Story = []
        styles = getSampleStyleSheet()
        styles.add(ParagraphStyle(name='Usuario', alignment=TA_RIGHT, fontName='Helvetica', fontSize=8))
        styles.add(ParagraphStyle(name='Titulo', alignment=TA_RIGHT, fontName='Helvetica', fontSize=18))
        styles.add(ParagraphStyle(name='Subtitulo', alignment=TA_RIGHT, fontName='Helvetica', fontSize=12))
        fecha = timezone.now().strftime('%Y-%m-%d')
        usuario = 'Usuario: ' + request.user.username + ' - Fecha: ' + datetime.datetime.now().strftime("%Y/%m/%d")
        Story.append(Paragraph(usuario, styles["Usuario"]))
        Story.append(Spacer(0, cm * 0.1))
        im0 = Image(settings.MEDIA_ROOT + '/imagenes/espacioPDF.png', width=490, height=3)
        im0.hAlign = 'CENTER'
        Story.append(im0)
        titulo = 'SISTEMA OBRAS PARTICULARES'
        Story.append(Paragraph(titulo, styles["Titulo"]))
        Story.append(Spacer(0, cm * 0.20))
        subtitulo = 'Listado de Profesionales activos'
        Story.append(Paragraph(subtitulo, styles["Subtitulo"]))
        Story.append(Spacer(0, cm * 0.15))
        Story.append(im0)
        Story.append(Spacer(0, cm * 0.5))
        encabezados = ('NOMBRE', 'APELLIDO', 'TELEFONO','PROFESION',
                       'MATRICULA','DOMICILIO','MAIL')
        tramites = Tramite.objects.all()  # puse con inspeccion solo para fines de mostrar algo
        personas = Profesional.objects.all()
        profesionales = []
        profesionales = Tramite.objects.select_related().values('profesional__persona__nombre',
                                                                'profesional__persona__apellido',
                                                                'profesional__persona__telefono',
                                                                'profesional__profesion',
                                                                'profesional__matricula',
                                                                'profesional__persona__domicilio',
                                                                'profesional__persona__mail').all().distinct()  # puse con inspeccion solo para fines de mostrar algo

        detalles = [(profesional['profesional__persona__nombre'],
                                                                profesional['profesional__persona__apellido'],
                                                                profesional['profesional__persona__telefono'],
                                                                profesional['profesional__profesion'],
                                                                profesional['profesional__matricula'],
                                                                profesional['profesional__persona__domicilio'],
                                                                profesional['profesional__persona__mail']) for
                    profesional in
                    profesionales
                    ]
        detalle_orden = Table([encabezados] + detalles, colWidths=[1.5 * cm, 1.8 * cm, 1.8 * cm, 3.3 * cm, 2 * cm,3 * cm,
                                                                   6 * cm])
        detalle_orden.setStyle(TableStyle(
            [
                ('ALIGN', (0, 0), (0, 0), 'CENTER'),
                ('GRID', (0, 0), (-1, -1), 1, colors.gray),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('LINEBELOW', (0, 0), (-1, 0), 2, colors.darkblue),
                ('BACKGROUND', (0, 0), (-1, 0), colors.dodgerblue),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ]
        ))
        detalle_orden.hAlign = 'CENTER'
        Story.append(detalle_orden)
        doc.build(Story)
        return response

#reporte tramites con final de obra solicitado

class ReporteSolicitudFinalObraExcel(TemplateView):
    def get(self, request, *args, **kwargs):
        tramites = Tramite.objects.en_estado(FinalObraSolicitado)
        wb = Workbook()
        ws = wb.active
        ws.merge_cells('B1:G1')
        ws.merge_cells('D2:F2')
        ws.merge_cells('G2:I2')
        ws.merge_cells('J2:K2')
        ws.merge_cells('L2:M2')
        ws['B1'] = 'REPORTE DE TRAMITES FINAL OBRA SOLICITADO'
        ws['B2'] = 'NUMERO'
        ws['C2'] = 'MEDIDAS'
        ws['D2'] = 'TIPO'
        ws['G2'] = 'ESTADO'
        ws['J2'] = 'PROFESIONAL'
        ws['L2'] = 'PROPIETARIO'
        cont = 3
        for tramite in tramites:
            ws.cell(row=cont, column=2).value = tramite.id
            ws.cell(row=cont, column=3).value = tramite.medidas
            ws.cell(row=cont, column=4).value = str(tramite.tipo_obra)
            ws.cell(row=cont, column=7).value = str(tramite.estado())
            ws.cell(row=cont, column=10).value = str(tramite.profesional)
            ws.cell(row=cont, column=12).value = str(tramite.propietario)
            cont = cont + 1
        nombre_archivo = "ReportePersonasExcel.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response

class ReporteSolicitudFinalObraPdf(View):
    def get(self, request, *args, **kwargs):
        filename = "Informe de tramites.pdf"
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="%s"' % filename
        doc = SimpleDocTemplate(
            response,
            pagesize=letter,
            rightMargin=72,
            leftMargin=72,
            topMargin=15,
            bottomMargin=28,
        )
        Story = []
        styles = getSampleStyleSheet()
        styles.add(ParagraphStyle(name='Usuario', alignment=TA_RIGHT, fontName='Helvetica', fontSize=8))
        styles.add(ParagraphStyle(name='Titulo', alignment=TA_RIGHT, fontName='Helvetica', fontSize=18))
        styles.add(ParagraphStyle(name='Subtitulo', alignment=TA_RIGHT, fontName='Helvetica', fontSize=12))
        fecha = timezone.now().strftime('%Y-%m-%d')
        usuario = 'Usuario: ' + request.user.username + ' - Fecha: ' + datetime.datetime.now().strftime("%Y/%m/%d")
        Story.append(Paragraph(usuario, styles["Usuario"]))
        Story.append(Spacer(0, cm * 0.1))
        im0 = Image(settings.MEDIA_ROOT + '/imagenes/espacioPDF.png', width=490, height=3)
        im0.hAlign = 'CENTER'
        Story.append(im0)
        titulo = 'SISTEMA OBRAS PARTICULARES'
        Story.append(Paragraph(titulo, styles["Titulo"]))
        Story.append(Spacer(0, cm * 0.20))
        subtitulo = 'Listado de tramites con solicitud final de obra'
        Story.append(Paragraph(subtitulo, styles["Subtitulo"]))
        Story.append(Spacer(0, cm * 0.15))
        Story.append(im0)
        Story.append(Spacer(0, cm * 0.5))
        encabezados = ('NUMERO', 'MEDIDAS', 'TIPO','ESTADO',
                       'PROFESIONAL','PROPIETARIO')
        detalles = [(tramite.id, tramite.medidas, tramite.tipo_obra, tramite.estado(),
                     tramite.profesional, tramite.propietario)
                    for
                    tramite in
                    Tramite.objects.en_estado(FinalObraSolicitado)]
        detalle_orden = Table([encabezados] + detalles, colWidths=[2 * cm, 2 * cm, 4 * cm, 4 * cm, 3 * cm,
                                                                   3 * cm])
        detalle_orden.setStyle(TableStyle(
            [
                ('ALIGN', (0, 0), (0, 0), 'CENTER'),
                ('GRID', (0, 0), (-1, -1), 1, colors.gray),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('LINEBELOW', (0, 0), (-1, 0), 2, colors.darkblue),
                ('BACKGROUND', (0, 0), (-1, 0), colors.dodgerblue),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ]
        ))
        detalle_orden.hAlign = 'CENTER'
        Story.append(detalle_orden)
        doc.build(Story)
        return response

def planilla_visado_impresa_administrativo(request, pk_tramite):
    planilla = get_object_or_404(PlanillaDeVisado, id=pk_tramite)
    tramite = get_object_or_404(Tramite, pk=planilla.tramite_id)
    filas = FilaDeVisado.objects.all()
    columnas = ColumnaDeVisado.objects.all()
    try:
        elementos = planilla.elementos.filter(activo=1)
        items = planilla.items.filter(activo=1)
        obs = planilla.observacion
        contexto = {'tramite': tramite,
                    'planilla': planilla,
                    'filas': filas,
                    'columnas': columnas,
                    'elementos': elementos,
                    'items': items,
                    'obs': obs,
                    }
        return render(request, 'persona/administrativo/planilla_de_visado_impresa_administrativo.html', contexto)
    except:
        contexto = {
            'tramite': tramite,
            'planilla': planilla,
            'filas': filas,
            'columnas': columnas,
            'obs': obs,
        }
        return render(request, 'persona/administrativo/planilla_de_visado_impresa_administrativo.html', contexto)


def planilla_inspeccion_impresa_administrativo(request, pk_tramite):
    planilla = get_object_or_404(PlanillaDeInspeccion, id=pk_tramite)
    tramite = get_object_or_404(Tramite, id=planilla.tramite_id)
    items = ItemInspeccion.objects.all()
    categorias = CategoriaInspeccion.objects.all()
    try:
        detalles = planilla.detalles.all()
        contexto = {
            'tramite': tramite,
            'planilla': planilla,
            'items': items,
            'categorias': categorias,
            'detalles': detalles,
        }
        return render(request, 'persona/administrativo/planilla_de_inspeccion_impresa_administrativo.html', contexto)

    except:
        contexto = {
            'tramite': tramite,
            'planilla': planilla,
            'items': items,
            'categorias': categorias,
        }
        return render(request, 'persona/administrativo/planilla_de_inspeccion_impresa_administrativo.html', contexto)


#-------------------------------------------------------------------------------------------------------------------
#visador -----------------------------------------------------------------------------------------------------------
from planilla_visado import forms as pforms
from planilla_visado import models as pmodels

@login_required(login_url="login")
@grupo_requerido('visador')
def mostrar_visador(request):
    contexto = {
        "ctxtramaceptado": tramites_aceptados(request),
        "ctxtramvisados": tramites_visados(request),
        "ctxmis_visados": mis_visados(request),
        "ctxtramvisadosnoaprobados": visados_noaproabados(request)
    }
    return render(request, 'persona/visador/visador.html', contexto)

def mis_visados(request):
    usuario = request.user
    estados = Estado.objects.all()
    tipo = 3 #visado
    tramites = Tramite.objects.en_estado(Visado)
    tramites_del_visador = filter(lambda t: t.estado().usuario == usuario, tramites)
    contexto = {"tramites_del_visador": tramites_del_visador}
    return tramites_del_visador

def tramites_aceptados(request):
    aceptados = Tramite.objects.en_estado(Aceptado)
    contexto = {'tramites': aceptados}
    return contexto

def tramites_visados(request):
    usuario = request.user
    estados = Estado.objects.all()
    tipo = 3 #visado
    argumentos = [Visado]
    tramites = Tramite.objects.en_estado(Visado)
    tramites_del_visador = filter(lambda t: t.estado().usuario == usuario, tramites)
    contexto = {"tramites_del_visador": tramites_del_visador}
    return contexto

def visados_noaproabados(request):
    usuario = request.user
    estados = Estado.objects.all()
    tipo = 4
    argumentos = [Corregido]
    tramites = Tramite.objects.en_estado(Corregido)
    tramites_del_visador = filter(lambda t: t.estado().usuario == usuario, tramites)
    contexto = {"tramites_del_visador": tramites_del_visador}
    return contexto

def ver_documentos_para_visado(request, pk_tramite):
    tipos_de_documentos_requeridos = TipoDocumento.get_tipos_documentos_para_momento(TipoDocumento.VISAR)
    FormularioDocumentoSet = FormularioDocumentoSetFactory(tipos_de_documentos_requeridos)
    inicial = metodo(tipos_de_documentos_requeridos)
    documento_set = FormularioDocumentoSet(initial=inicial)
    tramite = get_object_or_404(Tramite, pk=pk_tramite)
    return render(request, 'persona/visador/ver_documentos_tramite.html', {'tramite': tramite, 'documentos_requeridos': tipos_de_documentos_requeridos})

def ver_documentos_visados(request, pk_tramite):
    tramite = get_object_or_404(Tramite, pk=pk_tramite)
    planilla = get_object_or_404(PlanillaDeVisado, pk=tramite.id)
    items = planilla.items.all()
    filas = FilaDeVisado.objects.all()
    columnas = ColumnaDeVisado.objects.all()
    elementos = planilla.elementos.all()
    contexto = {
        'planilla':planilla,
        'filas': filas,
        'columnas': columnas,
        'items': items,
        'elementos': elementos,
    }
    return render(request, 'persona/visador/ver_documentos_visados.html', {'tramite': tramite, 'planilla':planilla, 'filas':filas, 'columnas':columnas, 'elementos':elementos, 'items':items})

from planilla_visado.models import FilaDeVisado, ColumnaDeVisado

def ver_planilla_visado(request):
    items = ItemDeVisado.objects.filter(activo=True)
    filas = FilaDeVisado.objects.all()
    columnas = ColumnaDeVisado.objects.all()
    elementos = Elemento_Balance_Superficie.objects.all()
    return render(request, 'persona/visador/ver_planilla_visado.html', {'tramite': tramite, 'items':items, 'filas':filas, 'columnas':columnas, 'elementos':elementos})

def generar_planilla_impresa(request, pk_tramite):
    tramite = get_object_or_404(Tramite, pk=pk_tramite)
    planillas = PlanillaDeVisado.objects.filter(tramite_id=tramite.id)  # busca las planillas que tengan el id del tramite
    filas = FilaDeVisado.objects.all()
    columnas = ColumnaDeVisado.objects.all()
    elementos = Elemento_Balance_Superficie.objects.all()
    try:
        if (len(planillas) > 1):
            aux = planillas[0]
            for p in planillas:
                if (p.id > aux.id):  # obtiene el ultimo visado del tramite
                    plan = p
                    aux=p
                else:
                    plan = aux
            planilla = PlanillaDeVisado.objects.get(id=plan.id)  # PlanillaDeVisado.objects.filter(tramite_id=tramite.id)# busca las planillas que tengan el id del tramite
        else:
            planilla = PlanillaDeVisado.objects.get(tramite_id=tramite.id)  # PlanillaDeVisado.objects.filter(tramite_id=tramite.id)# busca las planillas que tengan el id del tramite
        balance = planilla.elementos.all()
        items = planilla.items.all()
        obs = planilla.observacion
        contexto = {
            'tramite': tramite,
            'planilla': planilla,
            'filas': filas,
            'columnas': columnas,
            'items': items,
            'balance': balance,
            'elementos': elementos,
            'obs': obs,
        }
    except:
        contexto = {
            'tramite': tramite,
            'filas': filas,
            'columnas': columnas,
            'elementos': elementos
        }
    finally:
        return render(request, 'persona/visador/generar_planilla_impresa.html', contexto)


def mostrar_visados_noaprobados(request, pk_tramite):
    tramite = get_object_or_404(Tramite, pk=pk_tramite)
    planillas=PlanillaDeVisado.objects.filter(tramite_id=tramite.id)# busca las planillas que tengan el id del tramite
    if (len(planillas) > 1):
        aux = planillas[0]
        for p in planillas:
            if (p.id > aux.id):  #obtiene el ultimo visado del tramite
                plan = p
                aux = p
            else:
                plan= aux
        planilla = PlanillaDeVisado.objects.get(id=plan.id)  # PlanillaDeVisado.objects.filter(tramite_id=tramite.id)# busca las planillas que tengan el id del tramite
        obs = planilla.observacion
    else:
        planilla = PlanillaDeVisado.objects.get(tramite_id=pk_tramite)  # PlanillaDeVisado.objects.filter(tramite_id=tramite.id)# busca las planillas que tengan el id del tramite
    try:
        filas = FilaDeVisado.objects.all()
        columnas = ColumnaDeVisado.objects.all()
        obs = planilla.observacion
        elementos = planilla.elementos.all()
        items = planilla.items.all()
        obs = planilla.observacion
        contexto = {
            'tramite': tramite,
            'planilla': planilla,
            'filas': filas,
            'columnas': columnas,
            'items': items,
            'elementos': elementos,
            'obs': obs,
        }
    except:
        contexto = {
            'tramite': tramite,
            'planilla': planilla,
            'filas': filas,
            'columnas': columnas
        }
    finally:
        return render(request, 'persona/visador/mostrar_visador_noaprobados.html', contexto)

def cargar_planilla_visado(request,pk_tramite):
    tramite = get_object_or_404(Tramite, pk=pk_tramite)
    items = ItemDeVisado.objects.filter(activo=True)
    planilla=PlanillaDeVisado.objects.filter(tramite_id=tramite).last() #ultima planilla de visado
    p=[]
    el=[]
    aux=0
    elementos = Elemento_Balance_Superficie.objects.all()
    if planilla is not None:
        itemsPlanilla=planilla.items.all().distinct()
        elementosPlanilla=planilla.elementos.all()
        if itemsPlanilla is not None:
            for i in items:
                aux=filter(lambda iP: iP.columna_de_visado.nombre == i.columna_de_visado.nombre  and i.fila_de_visado.nombre == iP.fila_de_visado.nombre,itemsPlanilla)
                if len(aux)>=1:
                    b = [i,1]
                    p.append(b)
                else:
                    if len(aux)==0:
                        b = [i,0]
                        p.append(b)

        if elementosPlanilla is not None:
            for e in elementos:
                aux = filter(lambda eP: eP.nombre == e.nombre, elementosPlanilla)
                if len(aux) >= 1:
                    b = [e, 1]
                    el.append(b)
                else:
                    if len(aux) == 0:
                        b = [e, 0]
                        el.append(b)
    filas = FilaDeVisado.objects.all()
    columnas = ColumnaDeVisado.objects.all()
    try:
        monto="{0:.2f}".format(tramite.monto_a_pagar)
    except:
        monto=None
    return render(request, 'persona/visador/planilla_visado.html', {'tramite': tramite, 'items':items, 'filas':filas, 'columnas':columnas, 'elementos':elementos,'planilla':p, 'elementosPlanilla':el, "monto":monto})

def planilla_visado(request, pk_tramite):
    if request.method == "POST":
        observacion = request.POST["observaciones"]
        tram = request.POST['tram']
        monto_permiso = request.POST['monto']
        if "Envia Planilla de visado" in request.POST:
            no_aprobar_visado(request, tram, observacion,monto_permiso)
            print("no aprobar visado")
        else:
            if verificarItems(request, tram, monto_permiso):
                aprobar_visado(request, tram, monto_permiso)
            else:
                messages.add_message(request, messages.ERROR, "Para aprobar el visado debe ingresar 4 o mas items y dos o mas elemenetos de visado")

    return redirect('visador')


from planilla_visado.models import PlanillaDeVisado


def verificarItems(request, pk_tramite, monto_permiso):
    listItems = []
    listElementos = []
    listadoItems = []
    listadoElementos = []
    boleano = False
    tramite = get_object_or_404(Tramite, pk=pk_tramite)
    planilla = PlanillaDeVisado()
    planilla.tramite = tramite
    for name, value in request.POST.items():
        if name.startswith('item'):
            ipk = name.split('-')[1]
            listItems.append(ipk)
    items = ItemDeVisado.objects.filter(activo=True)
    for item in items:
        for i in listItems:
            if (item.id == int(i)):
                listadoItems.append(item)
    for name, value in request.POST.items():
        if name.startswith('elemento'):
            ipk = name.split('-')[1]
            listElementos.append(ipk)
    elementos = Elemento_Balance_Superficie.objects.filter(activo=True)
    for elemento in elementos:
        for i in listElementos:
            if (elemento.id == int(i)):
                listadoElementos.append(elemento)
    if len(listadoItems) >= 4 and len(listElementos) >= 2:
        boleano = True
    else:
        boleano = False
    return boleano


def aprobar_visado(request, pk_tramite, monto):
    list_items = []
    list_elementos = []
    tramite = get_object_or_404(Tramite, pk=pk_tramite)
    planilla = PlanillaDeVisado()
    planilla.tramite = tramite
    planilla.save()
    for name, value in request.POST.items():
        if name.startswith('item'):
            ipk= name.split('-')[1]
            list_items.append(ipk)
    items = ItemDeVisado.objects.filter(activo=True)
    for item in items:
        for i in list_items:
            if (item.id == int(i)):
                planilla.agregar_item(item)
    planilla.save()
    for name, value in request.POST.items():
        if name.startswith('elemento'):
            ipk= name.split('-')[1]
            list_elementos.append(ipk)
    elementos = Elemento_Balance_Superficie.objects.all()
    for elemento in elementos:
        for i in list_elementos:
            if (elemento.id == int(i)):
                planilla.agregar_elemento(elemento)
                planilla.save()
    planilla.save()
    usuario = request.user
    tramite.hacer(tramite.VISAR, usuario)
    tramite.monto_a_pagar= monto
    tramite.save()
    messages.add_message(request, messages.SUCCESS, 'Tramite visado aprobado')
    return redirect('visador')

def no_aprobar_visado(request, pk_tramite, observacion,monto):
    list_items = []
    list_elementos = []
    tramite = get_object_or_404(Tramite, pk=pk_tramite)
    planilla = PlanillaDeVisado()
    planilla.tramite = tramite
    planilla.observacion = observacion
    planilla.save()
    for name, value in request.POST.items():
        if name.startswith('item'):
            ipk= name.split('-')[1]
            list_items.append(ipk)
    items = ItemDeVisado.objects.filter(activo=True)
    for item in items:
        for i in list_items:
            if (item.id == int(i)):
                planilla.agregar_item(item)
    planilla.save()
    for name, value in request.POST.items():
        if name.startswith('elemento'):
            ipk= name.split('-')[1]
            list_elementos.append(ipk)
    elementos = Elemento_Balance_Superficie.objects.all()
    for elemento in elementos:
        for i in list_elementos:
            if (elemento.id == int(i)):
                planilla.agregar_elemento(elemento)
                planilla.save()
    planilla.save()
    usuario = request.user
    tramite = get_object_or_404(Tramite, pk=pk_tramite)
    tramite.monto_a_pagar= monto
    planilla.save()
    tramite.hacer(tramite.CORREGIR, usuario, observacion)
    tramite.save()
    messages.add_message(request, messages.SUCCESS, 'Tramite con visado no aprobado')
    return redirect('visador')

def tramites_visados_imprimible(request):
    usuario = request.user
    estados = Estado.objects.all()
    tipo = 3  # visado
    argumentos = [Visado]
    tramites = Tramite.objects.en_estado(Visado)
    tramites_del_visador = filter(lambda t: t.estado().usuario == usuario, tramites)
    contexto = {"tramites_del_visador": tramites_del_visador}
    return render(request, 'persona/visador/tramites_visados_imprimible.html', contexto)


class ReporteTramitesAceptadosExcel(TemplateView):

    def get(self, request, *args, **kwargs):
        tramites = Tramite.objects.en_estado(Aceptado)
        wb = Workbook()
        ws = wb.active
        ws['B1'] = 'REPORTE DE TRAMITES ACEPTADOS'
        ws.merge_cells('B1:F1')
        ws['C2'] = 'TIPO_DE_OBRA'
        ws['D2'] = 'PROFESIONAL'
        ws['E2'] = 'PROPIETARIO'
        ws['F2'] = 'MEDIDAS'
        cont = 3
        for tramite in tramites:
            ws.cell(row=cont, column=3).value = str(tramite.tipo_obra)
            ws.cell(row=cont, column=4).value = str(tramite.profesional)
            ws.cell(row=cont, column=5).value = str(tramite.propietario)
            ws.cell(row=cont, column=6).value = tramite.medidas
            cont = cont + 1
        nombre_archivo = "ReportePersonasExcel.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response

class ReporteTramitesAceptadosPdf(View):

    def get(self, request, *args, **kwargs):

        filename = "Informe de tramites Aceptados " + datetime.datetime.now().strftime("%d/%m/%Y") + ".pdf"
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="%s"' % filename
        doc = SimpleDocTemplate(
            response,
            pagesize=letter,
            rightMargin=72,
            leftMargin=72,
            topMargin=15,
            bottomMargin=28,
        )
        Story = []
        styles = getSampleStyleSheet()
        styles.add(ParagraphStyle(name='Usuario', alignment=TA_RIGHT, fontName='Helvetica', fontSize=8))
        styles.add(ParagraphStyle(name='Titulo', alignment=TA_RIGHT, fontName='Helvetica', fontSize=18))
        styles.add(ParagraphStyle(name='Subtitulo', alignment=TA_RIGHT, fontName='Helvetica', fontSize=12))
        usuario = 'Usuario: ' + request.user.username + ' - Fecha: ' + datetime.datetime.now().strftime("%Y/%m/%d")
        Story.append(Paragraph(usuario, styles["Usuario"]))
        Story.append(Spacer(0, cm * 0.15))
        im0 = Image(settings.MEDIA_ROOT + '/imagenes/espacioPDF.png', width=490, height=3)
        im0.hAlign = 'CENTER'
        Story.append(im0)
        titulo = 'SISTEMA OBRAS PARTICULARES'
        Story.append(Paragraph(titulo, styles["Titulo"]))
        Story.append(Spacer(0, cm * 0.20))
        subtitulo = 'Reporte De Tramites Iniciados para visar'
        Story.append(Paragraph(subtitulo, styles["Subtitulo"]))
        Story.append(Spacer(0, cm * 0.15))
        Story.append(im0)
        Story.append(Spacer(0, cm * 0.5))
        encabezados = ('TIPO DE OBRA', 'PROFESIONAL', 'PROPIETARIO', 'MEDIDAS', 'ESTADO')
        detalles = [(tramite.tipo_obra, tramite.profesional, tramite.propietario, tramite.medidas, tramite.estado()) for
                    tramite in
                    Tramite.objects.en_estado(Aceptado)]
        detalle_orden = Table([encabezados] + detalles, colWidths=[4 * cm, 4 * cm, 4 * cm, 3 * cm, 2 * cm])
        detalle_orden.setStyle(TableStyle(
            [
                ('ALIGN', (0, 0), (0, 0), 'CENTER'),
                ('GRID', (0, 0), (-1, -1), 1, colors.gray),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('LINEBELOW', (0, 0), (-1, 0), 2, colors.darkblue),
                ('BACKGROUND', (0, 0), (-1, 0), colors.dodgerblue),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ]
        ))
        detalle_orden.hAlign = 'CENTER'
        Story.append(detalle_orden)
        doc.build(Story)
        return response

  #  reporte_tramites_visados_pdf

class ReporteTramitesVisadosExcel(TemplateView):
    def get(self, request, *args, **kwargs):
        tramites = Tramite.objects.en_estado(Visado)
        wb = Workbook()
        ws = wb.active
        ws['B1'] = 'REPORTE DE TRAMITES VISADOS'
        ws.merge_cells('B1:F1')
        ws['B2'] = 'NRO'
        ws['C2'] = 'MEDIDAS'
        ws['D2'] = 'TIPO'
        ws['E2'] = 'PROFESIONAL'
        ws['F2'] = 'PROPIETARIO'
        ws['G2'] = 'DOMICILIO'
        cont = 3
        for tramite in tramites:
            ws.cell(row=cont, column=2).value = str(tramite.id)
            ws.cell(row=cont, column=3).value = str(tramite.medidas)
            ws.cell(row=cont, column=4).value = str(tramite.tipo_obra)
            ws.cell(row=cont, column=5).value = str(tramite.profesional)
            ws.cell(row=cont, column=6).value = str(tramite.propietario)
            ws.cell(row=cont, column=7).value = str(tramite.domicilio)
            cont = cont + 1
        nombre_archivo = "ReportePersonasExcel.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response

class ReporteTramitesVisadosPdf(View):
    def get(self, request, *args, **kwargs):
        filename = "Informe de tramites visados " + datetime.datetime.now().strftime("%d/%m/%Y") + " .pdf"
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="%s"' % filename
        doc = SimpleDocTemplate(
            response,
            pagesize=letter,
            rightMargin=72,
            leftMargin=72,
            topMargin=15,
            bottomMargin=28,
        )
        Story = []
        styles = getSampleStyleSheet()
        styles.add(ParagraphStyle(name='Usuario', alignment=TA_RIGHT, fontName='Helvetica', fontSize=8))
        styles.add(ParagraphStyle(name='Titulo', alignment=TA_RIGHT, fontName='Helvetica', fontSize=18))
        styles.add(ParagraphStyle(name='Subtitulo', alignment=TA_RIGHT, fontName='Helvetica', fontSize=12))
        usuario = 'Usuario: ' + request.user.username + ' - Fecha: ' + datetime.datetime.now().strftime("%Y/%m/%d")
        Story.append(Paragraph(usuario, styles["Usuario"]))
        Story.append(Spacer(0, cm * 0.15))
        im0 = Image(settings.MEDIA_ROOT + '/imagenes/espacioPDF.png', width=490, height=3)
        im0.hAlign = 'CENTER'
        Story.append(im0)
        titulo = 'SISTEMA OBRAS PARTICULARES'
        Story.append(Paragraph(titulo, styles["Titulo"]))
        Story.append(Spacer(0, cm * 0.20))
        subtitulo = 'Reporte De Tramites Visados'
        Story.append(Paragraph(subtitulo, styles["Subtitulo"]))
        Story.append(Spacer(0, cm * 0.15))
        Story.append(im0)
        Story.append(Spacer(0, cm * 0.5))
        encabezados = ('NRO', 'MEDIDAS', 'TIPO','PROFESIONAL', 'PROPIETARIO', 'DOMICILIO')
        detalles = [(tramite.id, tramite.medidas, tramite.tipo_obra, tramite.profesional, tramite.propietario, tramite.domicilio)
                    for
                    tramite in
                    Tramite.objects.en_estado(Visado)]
        detalle_orden = Table([encabezados] + detalles, colWidths=[2 * cm, 2 * cm, 3 * cm, 3 * cm, 4 * cm])
        detalle_orden.setStyle(TableStyle(
            [
                ('ALIGN', (0, 0), (0, 0), 'CENTER'),
                ('GRID', (0, 0), (-1, -1), 1, colors.gray),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('LINEBELOW', (0, 0), (-1, 0), 2, colors.darkblue),
                ('BACKGROUND', (0, 0), (-1, 0), colors.dodgerblue),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ]
        ))
        detalle_orden.hAlign = 'CENTER'
        Story.append(detalle_orden)
        doc.build(Story)
        return response

#-------------------------------------------------------------------------------------------------------------------
#inspector ---------------------------------------------------------------------------------------------------------
from django_user_agents.utils import get_user_agent
from planilla_inspeccion.models import *

@login_required(login_url="login")
@grupo_requerido('inspector')
def mostrar_inspector(request):
    if (request.user_agent.is_mobile): # returns True
        return redirect('movil_')
    contexto = {
        "ctxtramitesvisadosyconinspeccion": tramites_visados_y_con_inspeccion(request),
         "ctxtramitesinspeccionados": tramites_inspeccionados_por_inspector(request),
         "ctxtramitesagendados": tramites_agendados_por_inspector(request),
         "ctxtramis_inspecciones": mis_inspecciones(request),
         "ctxlistadomensual_inspector": listado_inspecciones_mensuales(request),
         "ctxlistado_inspector": listado_inspector_movil(request),
    }
    return render(request, 'persona/inspector/inspector.html', contexto)

def mis_inspecciones(request):
    usuario = request.user
    tipo = 6 #con inspeccion
    tramitesConInspeccion = Tramite.objects.en_estado(ConInspeccion)
    tramites = filter(lambda t: t.estado().usuario == usuario, tramitesConInspeccion)
    return tramites

def tramites_visados_y_con_inspeccion(request):
    argumentos = [Visado, ConInspeccion,FinalObraSolicitado]
    tramites = Tramite.objects.en_estado(argumentos)
    tram = []
    for t in tramites:
        planillas = PlanillaDeInspeccion.objects.filter(tramite_id=t.id).count()
        if planillas <3:
            tram.append(t)
    return tram

def tramites_inspeccionados_por_inspector(request):
    usuario = request.user
    estados = Estado.objects.all()
    tipo = 7 #7
    tramites = Tramite.objects.en_estado(ConInspeccion)
    # tramites_del_inspector = filter(lambda t: t.estado().usuario == usuario, tramites)
    estados_inspeccionados = filter(lambda t: (t.estado().usuario is not None and t.estado().usuario == usuario), tramites)
  #  contexto = {"tramites_del_inspector": tramites_del_inspector}
    return estados_inspeccionados

def tramites_agendados_por_inspector(request):
    usuario = request.user
    tipo = 5
    tramites = Tramite.objects.en_estado(Agendado)
    tramites_del_inspector = filter(lambda t: t.estado().usuario == usuario and t.estado().rol==2, tramites)
    contexto = {"tramites_del_inspector": tramites_del_inspector}
    return tramites_del_inspector

def agendar_tramite(request, pk_tramite):
    tramite = get_object_or_404(Tramite, pk=pk_tramite)
    usuario = request.user
    fecha = convertidor_de_fechas(request.GET["msg"])
    rol = request.GET["rol"]
    tramite.hacer(Tramite.AGENDAR, usuario=usuario, fecha_inspeccion=fecha,inspector=usuario,rol=2) #tramite, fecha_inspeccion, inspector=None rol=2 para inspector
    messages.add_message(request, messages.SUCCESS, "Inspeccion agendada")
    return redirect('inspector')

def cargar_inspeccion(request, pk_tramite):
    if request.method == "POST":
        if "Agendar" in request.POST:
            tramite = get_object_or_404(Tramite, pk=pk_tramite)
            list_detalles = []
            numeroPlanilla=PlanillaDeInspeccion.objects.filter(tramite_id=pk_tramite).count()
            print(numeroPlanilla)
            for name,value in request.POST.items():
                if name.startswith('detalle'):
                    ipk=name.split('-')[1]
                    detalle = DetalleDeItemInspeccion.objects.get(id=ipk)
                    list_detalles.append(detalle)
            if len(list_detalles)<2 and numeroPlanilla==0:
                messages.add_message(request, messages.ERROR, "Para realizar la inspeccion debe ingresar 2 o mas items de inspeccion")
            else:
                planilla = PlanillaDeInspeccion()
                planilla.tramite = tramite
                planilla.save()
                for detalle in list_detalles:
                    planilla.agregar_detalle(detalle)
                # planilla.tramite = tramite
                planilla.save()
                usuario = request.user
                tramite.hacer(tramite.INSPECCIONAR, usuario)
                tramite.save()
                messages.add_message(request,messages.SUCCESS,"Inspeccion cargada")
        else:
            messages.add_message(request,messages.ERROR,"No se cargo la inspeccion")
    return redirect('inspector')

def rechazar_inspeccion(request, pk_tramite):
    tramite = get_object_or_404(Tramite, pk=pk_tramite)
    tramite.hacer(Tramite.INSPECCIONAR, request.user)
    tramite.hacer(Tramite.CORREGIR, request.user, request.POST["observaciones"])  #request.POST["observaciones"]
    messages.add_message(request, messages.ERROR, 'Inspeccion rechazada')
    return redirect('inspector')

def aceptar_inspeccion(request, pk_tramite):
    tramite = get_object_or_404(Tramite, pk=pk_tramite)
    tramite.hacer(Tramite.INSPECCIONAR, request.user)
    messages.add_message(request, messages.SUCCESS, 'Inspeccion aprobada')
    return redirect('inspector')

def ver_documentos_tramite_inspector(request, pk_tramite):
    tramite = get_object_or_404(Tramite, pk=pk_tramite)
    planilla = get_object_or_404(PlanillaDeInspeccion, pk=pk_tramite)
    items = ItemInspeccion.objects.all()
    categorias = CategoriaInspeccion.objects.all()
    detalles = planilla.detalles.all()
    contexto = {'planilla': planilla,'items':items,'categorias':categorias,'detalles':detalles}
    return render(request, 'persona/inspector/documentos_del_estado_inspector.html', {"tramite": tramite,
                                                                                   'planilla': planilla, 'items': items,
                                                                                   'categorias': categorias,
                                                                                   'detalles': detalles})
def documentos_inspector_estado(request, pk_estado):
    estado = get_object_or_404(Estado, pk=pk_estado)
    fecha = estado.timestamp
    fecha_str = date.strftime(fecha, '%d/%m/%Y %H:%M')
    documentos = estado.tramite.documentos.all()
    documentos_fecha = filter(lambda e:(date.strftime(e.fecha, '%d/%m/%Y %H:%M') == fecha_str), documentos)
    if (estado.tipo >5):
        planilla = None
        for p in PlanillaDeInspeccion.objects.all():
            if (p.tramite.pk == estado.tramite.pk):
                planilla = p
        items = ItemInspeccion.objects.all()
        categorias = CategoriaInspeccion.objects.all()
        detalles = planilla.detalles.all()
        contexto = {'documentos_de_fecha': documentos_fecha, 'items': items, 'categorias': categorias, 'detalles': detalles, 'planilla':planilla}
    else:
        contexto = {'documentos_de_fecha': documentos_fecha}
    return render(request,'persona/inspector/documentos_del_estado_inspector.html', contexto)

def generar_planilla_impresa_inspeccion(request, pk_tramite):
    tramite = get_object_or_404(Tramite, pk=pk_tramite)
    planillas = PlanillaDeInspeccion.objects.filter(tramite_id=tramite.id) #busca las planillas que tengan el id del tramite
    if (len(planillas)>1):
        aux=planillas[0]
        for p in planillas:
            if (p.id>aux.id):
                plan=p
            else:
                plan=aux
        planilla=get_object_or_404(PlanillaDeInspeccion, id=plan.id)
    else:
        planilla=get_object_or_404(PlanillaDeInspeccion, tramite_id=tramite.id)
    items = ItemInspeccion.objects.all()
    categorias = CategoriaInspeccion.objects.all()
    try:
        detalles = planilla.detalles.all()
        contexto = {
            'tramite': tramite,
            'planilla': planilla,
            'items': items,
            'categorias': categorias,
            'detalles': detalles,
        }
        return render(request, 'persona/inspector/generar_planilla_impresa_inspeccion.html', contexto)

    except:
        contexto = {
            'tramite':tramite,
            'planilla': planilla,
            'items': items,
            'categorias': categorias,
        }
        return render(request, 'persona/inspector/generar_planilla_impresa_inspeccion.html', contexto)

#REPORTES INSPECTOR //DE TODOS LOS LISTADOS HICE REPORTES


class ReporteTramitesAgendarInspeccionExcel(TemplateView):
    def get(self, request, *args, **kwargs):
        tramites = Tramite.objects.all()
        tramites_aagendar = filter(lambda t: t.estado == 3 or t.estado == 6, tramites)
        wb = Workbook()
        ws = wb.active
        ws['B1'] = 'REPORTE DE TRAMITES VISADOS'
        ws.merge_cells('B1:F1')
        ws['B2'] = 'NRO'
        ws['C2'] = 'MEDIDAS'
        ws['D2'] = 'TIPO'
        ws['E2'] = 'PROFESIONAL'
        ws['F2'] = 'PROPIETARIO'
        ws['G2'] = 'DOMICILIO'
        cont = 3
        for tramite in tramites:
            ws.cell(row=cont, column=2).value = str(tramite.id)
            ws.cell(row=cont, column=3).value = str(tramite.medidas)
            ws.cell(row=cont, column=4).value = str(tramite.tipo_obra)
            ws.cell(row=cont, column=5).value = str(tramite.profesional)
            ws.cell(row=cont, column=6).value = str(tramite.propietario)
            ws.cell(row=cont, column=7).value = str(tramite.domicilio)
            cont = cont + 1
        nombre_archivo = "ReportePersonasExcel.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response

class ReporteTramitesAgendarInspeccionPdf(View):
    def get(self, request, *args, **kwargs):
        filename = "Informe de tramites.pdf"
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="%s"' % filename
        doc = SimpleDocTemplate(
            response,
            pagesize=letter,
            rightMargin=72,
            leftMargin=72,
            topMargin=15,
            bottomMargin=28,
        )
        Story = []
        styles = getSampleStyleSheet()
        styles.add(ParagraphStyle(name='Usuario', alignment=TA_RIGHT, fontName='Helvetica', fontSize=8))
        styles.add(ParagraphStyle(name='Titulo', alignment=TA_RIGHT, fontName='Helvetica', fontSize=18))
        styles.add(ParagraphStyle(name='Subtitulo', alignment=TA_RIGHT, fontName='Helvetica', fontSize=12))
        usuario = 'Usuario: ' + request.user.username + ' - Fecha: ' + datetime.datetime.now().strftime("%Y/%m/%d")
        Story.append(Paragraph(usuario, styles["Usuario"]))
        Story.append(Spacer(0, cm * 0.15))
        im0 = Image(settings.MEDIA_ROOT + '/imagenes/espacioPDF.png', width=490, height=3)
        im0.hAlign = 'CENTER'
        Story.append(im0)
        titulo = 'SISTEMA OBRAS PARTICULARES'
        Story.append(Paragraph(titulo, styles["Titulo"]))
        Story.append(Spacer(0, cm * 0.20))
        subtitulo = 'Reporte De Tramites Visados'
        Story.append(Paragraph(subtitulo, styles["Subtitulo"]))
        Story.append(Spacer(0, cm * 0.15))
        Story.append(im0)
        Story.append(Spacer(0, cm * 0.5))
        encabezados = ('NRO', 'MEDIDAS', 'TIPO','PROFESIONAL', 'PROPIETARIO', 'DOMICILIO')
        detalles = [(tramite.id, tramite.medidas, tramite.tipo_obra, tramite.profesional, tramite.propietario, tramite.domicilio)
                    for
                    tramite in
                    Tramite.objects.en_estado(Visado)]
        detalle_orden = Table([encabezados] + detalles, colWidths=[2 * cm, 2 * cm, 3 * cm, 3 * cm, 4 * cm])
        detalle_orden.setStyle(TableStyle(
            [
                ('ALIGN', (0, 0), (0, 0), 'CENTER'),
                ('GRID', (0, 0), (-1, -1), 1, colors.gray),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('LINEBELOW', (0, 0), (-1, 0), 2, colors.darkblue),
                ('BACKGROUND', (0, 0), (-1, 0), colors.dodgerblue),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ]
        ))
        detalle_orden.hAlign = 'CENTER'
        Story.append(detalle_orden)
        doc.build(Story)
        return response


#------------------------------------------------------------------------------------------------------------------
#jefeinspector ----------------------------------------------------------------------------------------------------

@login_required(login_url="login")
@grupo_requerido('jefeinspector')
def mostrar_jefe_inspector(request):
    if (request.user_agent.is_mobile):  # returns True
        return redirect('movil_jefe')
    contexto = {
        "ctxtramitesconinspeccion": tramite_con_inspecciones_list(request),
        "ctxtramitesagendados": tramites_agendados_por_inspector(request),
        "ctxlistadosinspecciones":listado_inspecciones(request),
        "ctxlistadosinspecciones_mensuales_ji":listado_inspecciones_mensuales_jefe_inspector(request),
    }
    return render(request, 'persona/jefe_inspector/jefe_inspector.html', contexto)

def tramite_con_inspecciones_list(request):
    argumentos=[ConInspeccion, FinalObraSolicitado]
    tramites = Tramite.objects.en_estado(argumentos)
    tram=[]
    for t in tramites:
        planillas = PlanillaDeInspeccion.objects.filter(tramite_id=t.id).count()
        if planillas>=3:
            tram.append(t)
    contexto = {'tramites': tram}
    return contexto

def agendar_inspeccion_final(request,pk_tramite):
    tramite = get_object_or_404(Tramite,pk=pk_tramite)
    fecha = convertidor_de_fechas(request.GET["msg"])
    tramite.hacer(Tramite.AGENDAR, usuario=request.user, fecha_inspeccion=fecha, inspector=request.user,rol=1) #para agendar jefe inspector rol=1
    messages.add_message(request, messages.SUCCESS, "Inspeccion agendada")
    return redirect('jefeinspector')

def inspeccion_final(request,pk_tramite):
   tramite = get_object_or_404(Tramite, pk=pk_tramite)
   planilla = PlanillaDeInspeccion.objects.select_related().filter(tramite_id=tramite).last()
   p = []
   detalles = DetalleDeItemInspeccion.objects.all()
   if planilla is not None:
       detallesPlanilla = planilla.detalles.all()
       if detallesPlanilla is not None:
           aux = 0
           for d in detalles:
               for i in detallesPlanilla:
                   if i.categoria_inspeccion.nombre == d.categoria_inspeccion.nombre and i.nombre == d.nombre:
                       b = [1, d]
                       p.append(b)
                       aux = 1
                       break;
               if aux == 0:
                   b = [0, d]
                   p.append(b)
               else:
                   aux = 0
   items = ItemInspeccion.objects.all()
   categorias = CategoriaInspeccion.objects.all()
   contexto = {"tramite": tramite, "items": items, "detalles": detalles, "categorias": categorias, "planilla": p}
   return render(request, 'persona/jefe_inspector/cargar_inspeccion_final.html', contexto)

def completar_inspeccion_final(request,pk_tramite):
    if "Guardar" in request.POST:
        tramite = get_object_or_404(Tramite, pk=pk_tramite)
        planilla = PlanillaDeInspeccion()
        planilla.tramite = tramite
        planilla.save()
        list_detalles=[]
        for name,value in request.POST.items():
            if name.startswith('detalle'):
                ipk=name.split('-')[1]
                list_detalles.append(ipk)
        detalles = DetalleDeItemInspeccion.objects.all()
        for detalle in detalles:
            for i in list_detalles:
                if (detalle.id == int(i)):
                    planilla.agregar_detalle(detalle)
        planilla.save()
        u = request.user
        try:
            tramite.hacer(Tramite.INSPECCIONAR, usuario=u)#ConInspeccion->Inspeccionado
            tramite.hacer(Tramite.INSPECCIONAR, usuario=u)  # ConInspeccion->Inspeccionado
            tramite.save()
            messages.add_message(request, messages.SUCCESS, 'Inspeccion Finalizada')
        except:
            messages.add_message(request, messages.WARNING, "La inspeccion ya fue cargada")
    return redirect('jefeinspector')

def aceptar_inspeccion_final(request,pk_tramite):
    u = request.user
    tramite.hacer(Tramite.INSPECCIONAR, usuario=u, inspector=u)#agendado->ConInspeccion
    tramite.hacer(Tramite.INSPECCIONAR, usuario=u)#ConInspeccion->Inspeccionado
    messages.add_message(request, messages.SUCCESS, 'Inspeccion Finalizada')
    return redirect('jefeinspector')

# ve la inspeccion de un tramite o inspecciones
def ver_inspecciones(request, pk_tramite):
    tramite = get_object_or_404(Tramite,pk=pk_tramite)
    inspecciones = []
    for p in PlanillaDeInspeccion.objects.all():
        if (p.tramite.pk == int(pk_tramite)):
            inspecciones.append(p)
    items = ItemInspeccion.objects.all()
    categorias = CategoriaInspeccion.objects.all()
    contexto = {
        'tramite': tramite,
        'inspecciones': inspecciones,
        'items': items,
        'categorias': categorias
    }
    return render(request, 'persona/jefe_inspector/vista_de_inspecciones.html',contexto)

def planilla_inspeccion_impresa_jefeinspector(request, pk_tramite):
    planilla = get_object_or_404(PlanillaDeInspeccion, id=pk_tramite)
    tramite=get_object_or_404(Tramite, id=planilla.tramite_id)
    items = ItemInspeccion.objects.all()
    categorias = CategoriaInspeccion.objects.all()
    try:
        detalles = planilla.detalles.all()
        contexto = {
            'tramite': tramite,
            'planilla': planilla,
            'items': items,
            'categorias': categorias,
            'detalles': detalles,
        }
        return render(request, 'persona/jefe_inspector/planilla_inspeccion_impresa_jefeinspector.html', contexto)

    except:
        contexto = {
            'tramite':tramite,
            'planilla': planilla,
            'items': items,
            'categorias': categorias,
        }
        return render(request, 'persona/jefe_inspector/planilla_inspeccion_impresa_jefeinspector.html', contexto)


def listado_inspecciones(request):
    tramites=Tramite.objects.en_estado(Agendado)
    tram=[]
    for t in tramites:
        if t.estado().rol==1:
          tram.append(t)
    contexto={'tramites':tram}
    return contexto

def listado_inspecciones_mensuales_jefe_inspector(request):
    year=datetime.date.today().year
    mes=datetime.date.today().month
    dia=datetime.date.today().day
    diaFinal=monthrange(year, mes)
    usuario = request.user
    tramites_del_inspector = Tramite.objects.en_estado(Agendado)
    tramites = filter(lambda t: t.estado().usuario == usuario  and t.estado().fecha.date()<=datetime.date(year,mes,diaFinal[1]) and t.estado().fecha.date()>=datetime.date(year,mes,dia), tramites_del_inspector)
    tram=[]
    for t in tramites:
        if t.estado().rol == 1:
            tram.append(t)
    contexto = {'tramites':tram}
    return contexto
#------------------------------------------------------------------------------------------------------------------
#director ---------------------------------------------------------------------------------------------------------
from planilla_visado import forms as pforms
from planilla_visado.forms import FormularioColumnaVisado
from planilla_visado import models as pmodels
from planilla_visado.models import *
from planilla_inspeccion.forms import FormularioCategoriaInspeccion
from planilla_inspeccion.forms import FormularioItemInspeccion
from planilla_inspeccion.forms import FormularioDetalleItem
from planilla_inspeccion.models import *
from planilla_inspeccion.models import CategoriaInspeccion, ItemInspeccion, DetalleDeItemInspeccion
from tipos.models import TipoObra, Tipo_Pago
from tipos.forms import FormularioTipoObra
from tipos.forms import FormularioTipoPago

@login_required(login_url="login")
@grupo_requerido('director')
def mostrar_director(request):
    usuario = request.user
    items = ItemInspeccion.objects.all()
    detalles = DetalleDeItemInspeccion.objects.all()
    categorias = CategoriaInspeccion.objects.all()
    filas = FilaDeVisado.objects.all()
    columnas = ColumnaDeVisado.objects.all()
    itemsVisados = ItemDeVisado.objects.filter(activo=True)
    elementos = Elemento_Balance_Superficie.objects.all()
    tiposPagos = Tipo_Pago.objects.all()
    tiposObras = TipoObra.objects.filter(activo=1)
    values = {"items":items, "categorias":categorias, "detalles":detalles, "filas": filas, "columnas":columnas, "itemsVisados":itemsVisados, "elementos":elementos, "ctxtramites_anuales":inspecciones_realizadas_durante_el_anio(request),
              "tiposPagos":tiposPagos, "tiposObras": tiposObras
}
    FORMS_DIRECTOR.update({(k.NAME, k.SUBMIT): k for k in [
        pforms.PlanillaDeVisadoFormFactory(pmodels.FilaDeVisado.objects.all(), pmodels.ColumnaDeVisado.objects.all()),
          ]})
    for form_name, submit_name in FORMS_DIRECTOR:

        KlassForm = FORMS_DIRECTOR[(form_name, submit_name)]
        if request.method == "POST" and submit_name in request.POST:
            _form = KlassForm(request.POST)
            if _form.is_valid():
                resultado=_form.save(commit=False)
                try:
                    if isinstance(resultado,ColumnaDeVisado):
                                columna=ColumnaDeVisado.objects.get(nombre=resultado.nombre)
                                columna.activo=1
                                columna.save()
                    elif isinstance(resultado,FilaDeVisado):
                                fila=FilaDeVisado.objects.get(nombre=resultado.nombre)
                                fila.activo=1
                                fila.save()
                    elif isinstance(resultado,Tipo_Pago):
                                tipoPago=Tipo_Pago.objects.get(nombre=resultado.nombre)
                                tipoPago.activo=1
                                tipoPago.save()
                    elif isinstance(resultado,Elemento_Balance_Superficie):#elemento de visado
                                elemento=Elemento_Balance_Superficie.objects.get(nombre=resultado.nombre)
                                elemento.activo=1
                                elemento.descripcion=resultado.descripcion
                                elemento.save()
                    elif isinstance(resultado, TipoObra):
                                tipoObra = TipoObra.objects.get(nombre=resultado.nombre)
                                tipoObra.activo = 1
                                tipoObra.descripcion=resultado.descripcion
                                tipoObra.categorias=resultado.categorias
                                tipoObra.save()
                    elif isinstance(resultado, CategoriaInspeccion):
                                categoria = CategoriaInspeccion.objects.get(nombre=resultado.nombre)
                                categoria.activo = 1
                                categoria.save()
                                items= DetalleDeItemInspeccion.objects.filter(categoria_inspeccion=categoria).update(activo=1)

                    elif isinstance(resultado, ItemInspeccion):
                                item = ItemInspeccion.objects.get(nombre=resultado.nombre)
                                item.activo = 1
                                item.save()
                                items= DetalleDeItemInspeccion.objects.filter(item_inspeccion=item).update(activo=1)

                    elif isinstance(resultado, DetalleDeItemInspeccion):
                                nombre=re.search(re.escape('<DetalleDeItemInspeccion:   ') + "(.*)" + re.escape(','), repr(resultado)).group(1).split(',')
                                nombre=nombre[0]
                                detalle = DetalleDeItemInspeccion.objects.get(nombre=nombre,categoria_inspeccion=resultado.categoria_inspeccion, item_inspeccion=resultado.item_inspeccion)
                                detalle.activo = 1
                                detalle.save()
                    else:
                         _form.save()
                except:
                    resultado.save()
                messages.add_message(request, messages.SUCCESS, "La accion solicitada ha sido ejecutada con exito")
                return redirect('director')
            else:
                values["submit_name"] = submit_name
                messages.add_message(request, messages.ERROR, "La accion solicitada no a podido ser ejecutada")
            values[form_name] = _form
        else:
            values[form_name] = KlassForm()
    return render(request, 'persona/director/director.html', values)

FORMS_DIRECTOR = {(k.NAME, k.SUBMIT): k for k in {
    FormularioTipoDocumento,
    FormularioUsuarioPersona,
# este formulario no se necesitaria, solo se dan de alta visador, inspector y administrativo
    FormularioTipoObra,
    FormularioTipoDocumento,
    FormularioTipoPago,
    pforms.FormularioColumnaVisado,
    pforms.FormularioFilaVisado,
    pforms.FormularioItemDeVisado,
    #pforms.PlanillaDeVisadoFormFactory(pmodels.FilaDeVisado.objects.all(), pmodels.ColumnaDeVisado.objects.all()),
    pforms.FormularioElementoBalanceSuperficie,
    FormularioCategoriaInspeccion,
    FormularioItemInspeccion,
    FormularioDetalleItem
}}

from planilla_inspeccion.forms import FormularioCategoriaInspeccionModificada
from planilla_inspeccion.forms import FormularioItemInspeccionModificado
from planilla_inspeccion.forms import FormularioDetalleItemModificado
from planilla_visado.forms import FormularioFilaVisadoModificada
from planilla_visado.forms import FormularioElementoBalanceSuperficieModificado
from planilla_visado.forms import FormularioColumnaVisadoModificada
from tipos.forms import FormularioTipoPagoModificado
from tipos.forms import FormularioTipoObraModificada

def listado_tiposPago(request):
    tiposPagos = Tipo_Pago.objects.all()
    return render(request, 'persona/director/listado_tiposPago.html', {'tiposPagos':tiposPagos})

def editar_tipoPago(request, pk_tipoPago):
    tipoPago = Tipo_Pago.objects.get(id=pk_tipoPago)
    if request.method == 'GET':
        form = FormularioTipoPagoModificado(instance=tipoPago)
    else:
        form = FormularioTipoPagoModificado(request.POST, instance=tipoPago)
        if form.is_valid():
            form.save()
            messages.add_message(request, messages.SUCCESS, "Tipo de pago modificado correctamente")
        else:
            messages.add_message(request, messages.ERROR, "El tipo de pago no pudo ser modificado, ya existe un tipo de pago con ese nombre")
        return redirect('director')
    return render(request, 'persona/director/editar_tipoPago.html', {'form':form})

def delete_tipo_pago(request, pk_tipoPago):
    tipoPago = Tipo_Pago.objects.get(id=pk_tipoPago)
    if request.method == 'POST':
        tipoPago.activo = False
        tipoPago.save()
        messages.add_message(request, messages.SUCCESS, "El tipo de pago ha sido eliminado correctamente")
        return redirect('director')
    return render(request, "persona/director/delete_tipo_pago.html", {'tipoPago':tipoPago})

def listado_tiposObras(request):
    tiposObras = TipoObra.objects.all()
    return render(request, 'persona/director/listado_tiposObras.html', {'tiposObras':tiposObras})

def edit_tipoObra(request, pk_tipoObra):
    tipoObra = TipoObra.objects.get(id=pk_tipoObra)
    if request.method == 'GET':
        form = FormularioTipoObraModificada(instance=tipoObra)
    else:
        try:
            form = FormularioTipoObraModificada(request.POST, instance=tipoObra)
            
            if form.is_valid():
                form.save()
                messages.add_message(request, messages.SUCCESS, "Tipo de obra modificada correctamente")
            else:
                messages.add_message(request, messages.ERROR, "El tipo de obra no pudo ser modificada, ya existe un tipo de obra con ese nombre")

        except:
                messages.add_message(request, messages.ERROR, "El tipo de obra no pudo ser modificada, ya existe un tipo de obra con ese nombre")
        return redirect('director')
    return render(request, 'persona/director/edit_tipoObra.html', {'form':form})

def delete_tipoObra(request, pk_tipoObra):
    tipoObra = TipoObra.objects.get(id=pk_tipoObra)
    if request.method == 'POST':
        tipoObra.activo = False
        tipoObra.save()
        messages.add_message(request, messages.SUCCESS, "El tipo de obra ha sido eliminada correctamente")
        return redirect('director')
    return render(request, "persona/director/delete_tipoObra.html", {'tipoObra':tipoObra})

def listado_filas_visado(request):
    filas = FilaDeVisado.objects.all()
    return render(request, 'persona/director/listado_filas_visado.html', {'filas':filas})

def edit_fila_visado(request, pk_fila):
    fila = FilaDeVisado.objects.get(id=pk_fila)
    if request.method == 'GET':
        form = FormularioFilaVisadoModificada(instance=fila)
    else:
        form = FormularioFilaVisadoModificada(request.POST, instance=fila)
        if form.is_valid():
            form.save()
            messages.add_message(request, messages.SUCCESS, "La fila de visado fue modificada correctamente")
        else:
            messages.add_message(request, messages.ERROR, "La fila de visado no pudo ser modificada, ya existe una fila con ese nombre")
        return redirect('director')
    return render(request, "persona/director/edit_fila_visado.html", {'form':form})

def delete_fila_visado(request, pk_fila):
    fila = FilaDeVisado.objects.get(id=pk_fila)
    if request.method == 'POST':
        fila.activo = False
        fila.save()
        messages.add_message(request, messages.SUCCESS, "La fila de visado ha sido eliminada correctamente")
        return redirect('director')
    return render(request, "persona/director/delete_fila_visado.html", {'fila':fila})

def listado_columnas_visado(request):
    columnas = ColumnaDeVisado.objects.all()
    return render(request, 'persona/director/listado_columnas_visado.html', {'columnas':columnas})

def edit_columna_visado(request, pk_columna):
    columna = ColumnaDeVisado.objects.get(id=pk_columna)
    if request.method == 'GET':
        form = FormularioColumnaVisadoModificada(instance=columna)
    else:
        form = FormularioColumnaVisadoModificada(request.POST, instance=columna)
        if form.is_valid():
            form.save()
            messages.add_message(request, messages.SUCCESS, "La columna de visado fue modificada correctamente")
        else:
            messages.add_message(request, messages.ERROR, "La columna de visado no pudo ser modificada, ya existe una columna con ese nombre")
        return redirect('director')
    return render(request, "persona/director/edit_columna_visado.html", {'form':form})

def delete_columna_visado(request, pk_columna):
    columna = ColumnaDeVisado.objects.get(id=pk_columna)
    if request.method == 'POST':
        columna.activo = False
        columna.save()
        messages.add_message(request, messages.SUCCESS, "La columna de visado ha sido eliminada correctamente")
        return redirect('director')
    return render(request, "persona/director/delete_columna_visado.html", {'columna':columna})

def listado_elementos_visado(request):
    elementos = Elemento_Balance_Superficie.objects.all()
    return render(request, 'persona/director/listado_elemento_visado.html', {'elementos':elementos})

def edit_elemento_visado(request, pk_elemento):
    elemento = Elemento_Balance_Superficie.objects.get(id=pk_elemento)
    if request.method == 'GET':
        form = FormularioElementoBalanceSuperficieModificado(instance=elemento)
    else:
        form = FormularioElementoBalanceSuperficieModificado(request.POST, instance=elemento)
        if form.is_valid():
            form.save()
            messages.add_message(request, messages.SUCCESS, "El elemento de visado fue modificado correctamente")
        else:
            messages.add_message(request, messages.ERROR, "El elemento de visado no pudo ser modificado, ya existe un elemento con ese nombre")
        return redirect('director')
    return render(request, "persona/director/edit_elemento_visado.html", {'form':form})

def delete_elemento_visado(request, pk_elemento):
    elemento = Elemento_Balance_Superficie.objects.get(id=pk_elemento)
    if request.method == 'POST':
        elemento.activo = False
        elemento.save()
        messages.add_message(request, messages.SUCCESS, "El elemento de visado ha sido eliminado correctamente")
        return redirect('director')
    return render(request, "persona/director/delete_elemento_visado.html", {'elemento':elemento})

def listado_item_inspeccion(request):
    items = ItemInspeccion.objects.all()
    return render(request, 'persona/director/listado_items_planilla_inspeccion.html', {'items':items})

def edit_item_inspeccion(request, pk_item):
    item = ItemInspeccion.objects.get(id=pk_item)
    if request.method == 'GET':
        form = FormularioItemInspeccionModificado(instance=item)
    else:
        form = FormularioItemInspeccionModificado(request.POST, instance=item)
        if form.is_valid():
            form.save()
            messages.add_message(request, messages.SUCCESS, "El item de inspeccion fue modificado correctamente")
        else:
            messages.add_message(request, messages.ERROR, "El item de inspeccion no pudo ser modificado, ya existe un item con ese nombre")
        return redirect('director')
    return render(request, "persona/director/edit_item_inspeccion.html", {'form':form})

def delete_item_inspeccion(request, pk_item):
    item = ItemInspeccion.objects.get(id=pk_item)
    detalles = DetalleDeItemInspeccion.objects.all()
    if request.method == 'POST':
        item.activo = False
        item.save()
        for d in detalles:
            if d.item_inspeccion == item:
                d.activo = False
                d.save()
        messages.add_message(request, messages.SUCCESS, "El item de inspeccion ha sido eliminado correctamente")
        return redirect('director')
    return render(request, "persona/director/delete_item_inspeccion.html", {'item':item})

def listado_categoria_inspeccion(request):
    categorias = CategoriaInspeccion.objects.all()
    return render(request, 'persona/director/listado_categorias_planilla_inspeccion.html', {'categorias':categorias})

def edit_categoria_inspeccion(request, pk_categoria):
    categoria = CategoriaInspeccion.objects.get(id=pk_categoria)
    if request.method == 'GET':
        form = FormularioCategoriaInspeccionModificada(instance=categoria)
    else:
        form = FormularioCategoriaInspeccionModificada(request.POST, instance=categoria)
        if form.is_valid():
            form.save()
            messages.add_message(request, messages.SUCCESS, "La categoria de inspeccion fue modificada correctamente")
        else:
            messages.add_message(request, messages.ERROR, "La categoria de inspeccion no pudo ser modificada, ya existe una categoria con ese nombre")
        return redirect('director')
    return render(request, "persona/director/edit_categoria_inspeccion.html", {'form':form})

def delete_categoria_inspeccion(request, pk_categoria):
    categoria = CategoriaInspeccion.objects.get(id=pk_categoria)
    detalles = DetalleDeItemInspeccion.objects.all()
    if request.method == 'POST':
        categoria.activo = False
        categoria.save()
        for d in detalles:
            if d.categoria_inspeccion == categoria:
                d.activo = False
                d.save()
        messages.add_message(request, messages.SUCCESS, "La categoria de inspeccion ha sido eliminada correctamente")
        return redirect('director')
    return render(request, "persona/director/delete_categoria_inspeccion.html", {'categoria':categoria})

def listado_detalle_inspeccion(request):
    detalles = DetalleDeItemInspeccion.objects.all()
    return render(request, 'persona/director/listado_detalles_planilla_inspeccion.html', {'detalles':detalles})

def edit_detalle_inspeccion(request, pk_detalle):
    detalle = DetalleDeItemInspeccion.objects.get(id=pk_detalle)
    if request.method == 'GET':
        form = FormularioDetalleItemModificado(instance=detalle)
    else:
        form = FormularioDetalleItemModificado(request.POST, instance=detalle)
        if form.is_valid():
            form.save()
            messages.add_message(request, messages.SUCCESS, "El detalle de inspeccion fue modificado correctamente")
        else:
            messages.add_message(request, messages.ERROR, "El detalle de inspeccion no pudo ser modificado, ya existe un detalle con ese nombre")
        return redirect('director')
    return render(request, "persona/director/edit_detalle_inspeccion.html", {'form':form})

def delete_detalle_inspeccion(request, pk_detalle):
    detalle = DetalleDeItemInspeccion.objects.get(id=pk_detalle)
    if request.method == 'POST':
        detalle.activo = False
        detalle.save()
        messages.add_message(request, messages.SUCCESS, "El detalle de inspeccion ha sido eliminado correctamente")
        return redirect('director')
    return render(request, "persona/director/delete_detalle_inspeccion.html", {'detalle':detalle})

def cambiar_usuario_de_grupo(request):
    contexto = {
        "ctxempleados": empleados(request),
        "ctxgrupos": grupos(request),
    }
    return render(request, 'persona/director/cambiar_usuario_de_grupo.html', contexto)

def empleados(request):
    personas = Persona.objects.all()
    contexto = {'persona': personas}
    return contexto

def grupos(request):
    grupos = Group.objects.all()
    contexto = {'grupo': grupos}
    return contexto

def ver_listado_todos_tramites(request):
    argumentos = [Iniciado, Aceptado, Visado, Corregido, Agendado, ConInspeccion, Inspeccionado, FinalObraSolicitado]
    tramites = Tramite.objects.en_estado(argumentos)
    estados = []
    for t in tramites:
        estados.append(t.estado().tipo);
    estados_cant = dict(collections.Counter(estados))
    for n in range(1, 9):
        if (not estados_cant.has_key(n)):
            estados_cant.setdefault(n, 0);
    estados_datos = estados_cant.values()
    contexto = {'todos_los_tramites': tramites, "datos_estados":estados_datos, "label_estados":argumentos}
    return render(request, 'persona/director/vista_de_tramites.html', contexto)

def ver_listado_todos_usuarios(request):
    grupos = Group.objects.all()
    label_grupos = []
    for g in grupos:
        label_grupos.append(g.name)
    usuarios = Usuario.objects.all()
    cant_usuarios_grupos = []
    for u in usuarios:
        for gu in u.get_view_groups():
            cant_usuarios_grupos.append(gu)
    total_usuarios_grupos = dict(collections.Counter(cant_usuarios_grupos))
    for lg in grupos:
        if (not total_usuarios_grupos.has_key(lg)):
            total_usuarios_grupos.setdefault(lg, 0)
    datos_grupos = total_usuarios_grupos.values()
    return render(request, 'persona/director/vista_de_usuarios.html', {"label_grupos":label_grupos, "datos_grupos":datos_grupos})

def ver_todos_tramites(request):
    argumentos = [Iniciado, Aceptado, Visado, Corregido, Agendado, ConInspeccion, Inspeccionado, FinalObraSolicitado, Finalizado]
    tramites = Tramite.objects.en_estado(argumentos)
    estados = []
    for t in tramites:
        estados.append( { t.estado().timestamp.year : { str(t.estado().tipo):1 }})
    contador = defaultdict(collections.Counter)
    for dd in estados:
            for k, d in dd.items():
                    contador[k].update(d)
    contador.default_factory = None
    datosJSON = json.dumps(contador)
    contexto = {'todos_los_tramites': tramites, "label_estados":["Iniciado", "Aceptado","Visado", "Corregido", "Agendado", "Con Inspeccion", "Inspeccionado", "Final Obra Solicitado","Finalizado"],"datos":datosJSON}
    return render(request, 'persona/director/vista_de_todos_tramites.html', contexto)

def ver_tipos_de_obras_mas_frecuentes(request):
    tramites = Tramite.objects.all()
    tipos_obras = TipoObra.objects.filter(activo=1)
    total=tramites.count()
    list = []
    porcentaje = []
    datos=[]
    nombres=[]
    for t in tipos_obras:
        cant = Tramite.objects.filter(tipo_obra_id=t.id).exclude(tipo_obra_id__isnull=True).count()
        l = [t, cant]
        list.append(l)
        if cant!=0:
            datos.append(cant)
            nombres.append(t.nombre)
            porcentaje.append("{0:.2f}".format((cant / float(total)) * 100))
        cant=None
    aux = [nombres[i] + " " + porcentaje[i] + "%" for i in range(0, len(nombres))]
    titulo = "Tipos de obras mas frecuentes"
    grafico = pie_chart_with_legend(datos, aux, titulo)
    imagen = base64.b64encode(grafico.asString("png"))
    contexto = {"tipos_obras": list, "grafico":imagen}
    return render(request,'persona/director/tipos_de_obras_mas_frecuentes.html',contexto)

def add_legend(draw_obj, chart, data):
    legend = Legend()
    legend.alignment = 'right'
    legend.x = 10
    legend.y = 95
    legend.deltax = 3
    legend.dxTextSpace = 3
    legend.columnMaximum = 5
    legend.colorNamePairs = Auto(obj=chart)
    draw_obj.add(legend)

def pie_chart_with_legend(datos, nombres,titulo):
    drawing = Drawing(width=600, height=300)
    pie = Pie()
    pie.sideLabels = True
    pie.x = 300
    pie.y = 170
    pie.data = datos
    longitud=len(datos)
    col=colores(longitud,6)
    if (col is not None):
        for i in range(longitud): pie.slices[i].fillColor = col[i]
    pie.slices.popout = 8
    pie.labels = [cat for cat in nombres]
    pie.slices.strokeWidth = 0.5
    pie.slices.popout = 5
    drawing.add(pie)
    add_legend(drawing, pie, datos)
    return drawing

def ver_categorias_mas_frecuentes(request):
    nombres=[]
    porcentaje=[]
    planillas = PlanillaDeInspeccion.objects.all()
    tramites = Tramite.objects.all()
    tipos_categorias = CategoriaInspeccion.objects.filter(activo=1)
    d = DetalleDeItemInspeccion.objects.filter(activo=1)
    #total=d.count()
    
    list = []
    list_categorias = []
    datos = []
    total=0
    for t in tipos_categorias:
        cant = PlanillaDeInspeccion.objects.filter(detalles__categoria_inspeccion_id=t.id).exclude(detalles__categoria_inspeccion__activo=0).count()
        total+=cant
        l = [t, cant]
        list.append(l)
        if cant != 0:
            datos.append(cant)
            nombres.append(t.nombre)
        cant = None

    for l in list:
        porcentaje.append("{0:.2f}".format((l[1] / float(total)) * 100))
        # categorias = dict(collections.Counter(list_categorias))
    for cat in tipos_categorias:
         if cat.id in list:
             nombres.append(cat.nombre)
    aux=[nombres[i]+" "+porcentaje[i]+"%" for i in range( 0, len(nombres))]
    titulo = "Categorias mas frecuentes"
    grafico = pie_chart_with_legend(datos, aux, titulo)
    imagen = base64.b64encode(grafico.asString("png"))
    contexto = {
        "tipos_categorias": tipos_categorias,
        "detalles": d,
        "grafico": imagen,
    }
    return render(request, 'persona/director/categorias_mas_frecuentes.html', contexto)

def ver_profesionales_mas_requeridos(request):
    tramites_inspeccionados = Tramite.objects.en_estado(ConInspeccion) #aca deberia ir estado Finalizado
    profesionales = Profesional.objects.all()
    list = []
    list_profesionales = []
    datos=[]
    nombres=[]
    total=0
    porcentaje=[]
    for p in profesionales:
        cant=Tramite.objects.filter(profesional_id=p.id).exclude(profesional_id__isnull=True).count()
        m=[p,cant]
        if (m[1]!=0):
            list.append(m)
        if cant!=0:
            datos.append(cant)
            nombres.append(p.persona.nombre+" "+p.persona.apellido)
        total+=cant
        cant=None

    for d in datos:
        porcentaje.append("{0:.2f}".format((d / float(total)) * 100))
    aux = [nombres[i] + " " + porcentaje[i] + "%" for i in range(0, len(nombres))]
    titulo = "Profesionales mas requeridos"
    grafico = pie_chart_with_legend(datos, aux , titulo)
    imagen = base64.b64encode(grafico.asString("png"))
    contexto = {
        "profesionales": list,
        "grafico":imagen
    }
    return render(request, 'persona/director/profesionales_mas_requeridos.html',contexto)

def ver_barra_materiales(request):
    items = ItemInspeccion.objects.all()
    planillas = PlanillaDeInspeccion.objects.all()
    contexto = {
        "items":items,
    }
    if "Guardar" in request.POST:
        for name, value in request.POST.items():
            if name.startswith('item'):
                contexto = __busco_item__(value)
                titulo = "Materiales mas utilizados"
                grafico = pie_chart_with_legend(contexto["datos"], contexto["nombres"],titulo)
                imagen = base64.b64encode(grafico.asString("png"))
                return render(request, 'persona/director/materiales_mas_usados.html',{"datos":contexto,"tipo_item":value,"grafico":imagen})
    if "Volver" in request.POST:
        pass
    return render(request,'persona/director/barra_materiales.html',contexto)

def __busco_item__(item):
    i = get_object_or_404(ItemInspeccion, nombre=item)
    list = []
    detalles=[]
    datos=[]
    nombres=[]
    porcentaje=[]
    total=0
    detallesItems=DetalleDeItemInspeccion.objects.select_related().filter(item_inspeccion_id=i.id, activo=True)
    for i in range(0,len(detallesItems)):
        aux={'nombre':detallesItems[i].nombre,'categoria': detallesItems[i].categoria_inspeccion.nombre}
        detalles.append(aux)

    for n in range(0,len(detalles)):
        planillas = PlanillaDeInspeccion.objects.filter(detalles__in=detallesItems, detalles__nombre=detalles[n]['nombre'],detalles__categoria_inspeccion__nombre=detalles[n]['categoria'])
        cant=planillas.count()
        total+=cant
        m={'cantidad':cant,'nombre':detalles[n]['nombre'],'categoria':detalles[n]['categoria']}
        list.append(m)
    list.sort(reverse=True)
    if len(list)>10:
        tope=10
    else:
        tope=len(list)
    for i in range (0,tope):
        datos.append(list[i]['cantidad'])
        nombres.append(list[i]['nombre'])
    for d in datos:
        porcentaje.append("{0:.2f}".format((d / float(total)) * 100))
    aux = [nombres[i] + " " + porcentaje[i] + "%" for i in range(0, len(nombres))]
    contexto={'datos':datos,'nombres':aux,'detalles':list[0:10]}
    return contexto

def detalle_de_tramite(request, pk_tramite):
    tramite = get_object_or_404(Tramite, pk=pk_tramite)
    contexto0 = {'tramite': tramite}
    pk = int(pk_tramite)
    estados = tramite.estadosTramite()
   # estados_de_tramite = filter(lambda e: (e.tramite.pk == pk), estados)
    contexto1 = {'estados_del_tramite': estados}
    fechas_del_estado = [];
    for est in estados:
        fechas_del_estado.append(est.timestamp.strftime("%d/%m/%Y"));
    return render(request, 'persona/director/detalle_de_tramite.html', {"tramite": contexto0, "estados": contexto1, "fecha": fechas_del_estado})


'''
def documentos_del_estado(request, pk_estado):
    estado = get_object_or_404(Estado, pk=pk_estado)
    fecha = estado.timestamp
    fecha_str = date.strftime(fecha, '%d/%m/%Y %H:%M')
    documentos = estado.tramite.documentos.all()
    documentos_fecha = filter(lambda e: (date.strftime(e.fecha, '%d/%m/%Y %H:%M') <= fecha_str), documentos)
    contexto = {'documentos_de_fecha': documentos_fecha}
    planillas = []
    inspecciones = []
    documento = estado.tramite.documentacion_para_estado(estado)
    if (estado.tipo==1 or estado.tipo==2):
        contexto={'documentos':documentos_fecha}
    if (estado.tipo > 2 and estado.tipo < 5):
        for p in PlanillaDeVisado.objects.all():
            if (p.tramite.pk == estado.tramite.pk):
                planillas.append(p)
        filas = FilaDeVisado.objects.all()
        columnas = ColumnaDeVisado.objects.all()
        contexto = {
            'planillas': planillas,
            'filas': filas,
            'columnas': columnas,
        }
    if (estado.tipo >= 5 and estado.tipo <= 9):
        for p in PlanillaDeInspeccion.objects.all():
            if (p.tramite.pk == estado.tramite.pk):
                inspecciones.append(p)
        items = ItemInspeccion.objects.all()
        categorias = CategoriaInspeccion.objects.all()
        contexto = {
            'inspecciones': inspecciones,
            'items': items,
            'categorias': categorias,
        }
    return render(request, 'persona/director/documentos_del_estado.html', contexto)
'''

def documentos_del_estado(request, pk_estado):
    estado = get_object_or_404(Estado, pk=pk_estado)
    documentos = estado.tramite.documentacion_para_estado(estado)
    return render(request, 'persona/director/documentos_del_estado.html', documentos)

def generar_planilla_visado(request):
    filas = FilaDeVisado.objects.all()
    columnas = ColumnaDeVisado.objects.all()
    contexto = {'filas': filas}
    contexto_columnas = {'columnas': columnas}
    balancesSuperficies = Elemento_Balance_Superficie.objects.all()
    itemsVisados = ItemDeVisado.objects.filter(activo=True)
    contexto = {'filas': filas, 'columnas':columnas, 'itemsVisados':itemsVisados,'balancesSuperficies':balancesSuperficies}
    return render(request, 'persona/director/item_visado.html', contexto)

def ver_planilla_inspeccion(request):
     items = ItemInspeccion.objects.all()
     detalles = DetalleDeItemInspeccion.objects.all()
     categorias = CategoriaInspeccion.objects.all()
     contexto = {'items': items}
     return render(request, 'persona/director/ver_planilla_inspeccion.html', contexto)


def planilla_visado_impresa_director(request, pk_tramite):
    itemsV = ItemDeVisado.objects.filter(activo=True)
    lista=[]
    planilla = get_object_or_404(PlanillaDeVisado,id=pk_tramite)
    tramite = get_object_or_404(Tramite, pk=planilla.tramite_id)
    filas = FilaDeVisado.objects.all()
    columnas = ColumnaDeVisado.objects.all()
    try:
        elementos = planilla.elementos.filter(activo=1)
        items = planilla.items.filter(activo= 1)
        obs = planilla.observacion
        contexto={'tramite': tramite,
                  'planilla': planilla,
                  'filas': filas,
                  'columnas': columnas,
                  'elementos': elementos,
                  'items': items,
                  'obs': obs,
                  }
        return render(request, 'persona/director/planilla_visado_impresa_director.html',contexto)
    except:
         contexto = {
             'tramite': tramite,
             'planilla': planilla,
             'filas': filas,
             'columnas': columnas,
             'obs': obs,
         }
         return render(request, 'persona/director/planilla_visado_impresa_director.html', contexto)

def planilla_inspeccion_impresa_director(request, pk_tramite):
    planilla = get_object_or_404(PlanillaDeInspeccion, id=pk_tramite)
    tramite=get_object_or_404(Tramite, id=planilla.tramite_id)
    items = ItemInspeccion.objects.all()
    categorias = CategoriaInspeccion.objects.all()
    try:
        detalles = planilla.detalles.all()
        contexto = {
            'tramite': tramite,
            'planilla': planilla,
            'items': items,
            'categorias': categorias,
            'detalles': detalles,
        }
        return render(request, 'persona/director/planilla_inspeccion_impresa_director.html', contexto)

    except:
        contexto = {
            'tramite':tramite,
            'planilla': planilla,
            'items': items,
            'categorias': categorias,
        }
        return render(request, 'persona/director/planilla_inspeccion_impresa_director.html', contexto)

def ver_filtro_obra_fechas(request):
    listado_tramites = []
    list_estados_fechas = []
    porcentaje=[]
    total=0
    if "Guardar" in request.POST:
        tipos = TipoObra.objects.filter(activo=1)
        tramites = Tramite.objects.all()
        estados = Estado.objects.all()
        list = []
        list_obras = []
        for name, value in request.POST.items():
             if name == 'fechaInicial':
                 fechaInicial = value
             if name == 'fechaFinal':
                 fechaFinal = value
        datos = []
        nombres = []
        tipos = TipoObra.objects.filter(activo=True)
        if fechaInicial and fechaFinal is not None:
            rango = True
            e=Estado.objects.filter(timestamp__range=(fechaInicial,fechaFinal))
        else:
            rango = False
            e = Estado.objects.all()
        for c in tipos:
                estado = e.values_list('tramite', flat="True").distinct()  # filter(lambda r:(r.tramite.tipo_obra_id==3),e)
                tramites = Tramite.objects.filter(id__in=estado, tipo_obra_id=c.id, tipo_obra__activo=True)

                cant = len(tramites)
                if cant != 0:
                    datos.append(cant)
                    l = [c.nombre, cant]
                    nombres.append(c.nombre)
                    total+=cant
                    list.append(l)
                estados = None
                tramites = None
        for d in datos:
            porcentaje.append("{0:.2f}".format((d / float(total)) * 100))
        aux = [nombres[i] + " " + porcentaje[i] + "%" for i in range(0, len(nombres))]
        titulo = "Tipos de obras en rango seleccionado"
        if len(datos)>0:
         grafico = pie_chart_with_legend(datos, aux, titulo)
         imagen = base64.b64encode(grafico.asString("png"))
         contexto = {"tipos_obras":list,"grafico":imagen}#"tipos_obras": list}
        else:
         contexto = {"tipos_obras":list}#"tipos_obras": list}
        return render(request, 'persona/director/tipos_obras_periodo_fechas.html', contexto)
    else:
        pass
    return render(request,'persona/director/filtro_obra_fechas.html')

def colores(longitud,tope):
    if longitud > tope:
        colores = [
            PCMYKColor(90, 30, 0, 10, alpha=100), PCMYKColor(5, 80, 37, 30, alpha=100),
            PCMYKColor(0, 16, 18, 10, alpha=100),
            PCMYKColor(3, 60, 80, 10, alpha=100), PCMYKColor(0, 90, 20, 10, alpha=100),
            PCMYKColor(60, 0, 70, 10, alpha=100), PCMYKColor(0, 0, 50, 30, alpha=100),
            PCMYKColor(0, 0, 100, 10, alpha=100),
            PCMYKColor(27, 8, 15, 20, alpha=100), PCMYKColor(21, 0, 34, 10, alpha=100),
            PCMYKColor(0, 88, 37, 15, alpha=100),
            PCMYKColor(54, 10, 0, 70, alpha=100),
            PCMYKColor(0, 32, 0, 60, alpha=100),
            PCMYKColor(20, 40, 0, 25, alpha=100),
            PCMYKColor(0, 40, 83, 5, alpha=100),
            PCMYKColor(40, 20, 0, 10, alpha=100), PCMYKColor(30, 0, 0, 12, alpha=100),
            PCMYKColor(100, 67, 0, 23, alpha=100), PCMYKColor(70, 46, 0, 16, alpha=100),
            PCMYKColor(50, 33, 0, 11, alpha=100), PCMYKColor(80, 5, 0, 5, alpha=100),
            PCMYKColor(20, 13, 0, 4, alpha=100), PCMYKColor(10, 7, 0, 30, alpha=100),
            PCMYKColor(0, 78, 0,0, alpha=100),
            PCMYKColor(30, 80, 70, 20, alpha=100),
            PCMYKColor(120, 50, 0, 50, alpha=100), PCMYKColor(0, 0, 78, 30, alpha=100),
            PCMYKColor(0, 45, 0, 20, alpha=100), PCMYKColor(68, 0, 80, 10, alpha=100),
            PCMYKColor(20, 19, 80, 15, alpha=100)]
        return colores

def grafico_de_barras_v(datos,nombres, titulo,series):
    drawing = Drawing(width=530, height=230)
    my_title = String(280, 280, titulo, fontSize=18)
    bc = VerticalBarChart()
    longitud=len(series)
    col=colores(longitud,3)
    bc.data=datos
    bc.x = 50
    bc.y = 130
    bc.width=410
    bc.height=90
    bc.valueAxis.valueMin = 0
    lista=list(chain.from_iterable(datos))
    tope=max(lista)
    if tope > 0 and math.log10(tope)>1:
        bc.valueAxis.valueMax = tope+(tope/10)
        bc.groupSpacing = tope/10
        if max(lista) >1000000:
            bc.barWidth = 100000
    else:
        if max(lista) <10:
            if max(lista) <5:
                bc.valueAxis.valueMax = 5
                bc.valueAxis.valueStep = 2
            else:
                bc.valueAxis.valueMax = 15
                bc.groupSpacing = 10
                bc.valueAxis.valueStep = 5
        else:
            bc.valueAxis.valueMax = 50
            bc.groupSpacing = 10
        bc.barSpacing = 1.5
    bc.categoryAxis.labels.boxAnchor = 'ne'
    bc.categoryAxis.labels.dy = -2
    bc.categoryAxis.labels.angle = 30
    if (col is not None):
        for i in range(longitud): bc.bars[i].fillColor = col[i]
    for i in range(len(series)): bc.bars[i].name = series[i]
    bc.categoryAxis.categoryNames=[n for n in nombres]
    drawing.add(bc)
    drawing.add(my_title)
    add_legend(drawing, bc, datos)

    return drawing

def grafico_de_barras(datos,nombres,titulo):
    lc = HorizontalLineChart()
    drawing = Drawing(width=400, height=200)
    my_title = String(150, 180, titulo, fontSize=18)
    lc = HorizontalLineChart()
    lc.data=datos
    lc.categoryAxis.categoryNames=[n for n in nombres]
    lc.categoryAxis.labels.boxAnchor = 'n'
    lc.valueAxis.valueMin = 0
    lc.valueAxis.valueMax = 50
    lc.valueAxis.valueStep = 10
    drawing.add(lc)
    drawing.add(my_title)
    return drawing

def seleccionar_fecha_item_inspeccion(request):
    datos = []
    series = []
    nombres = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre",
               "Noviembre", "Diciembre"]
    item = []
    year=date.today().year
    cant=[]
    if "Guardar" in request.POST:
        listaItems=request.POST.getlist('item')
        total=DetalleDeItemInspeccion.objects.all().count()
        lista=[]
        totalItems=0
        subtotales=[]
        for l in listaItems:

            i=DetalleDeItemInspeccion.objects.get(id=str(l))
            item.append(i)
        for name, value in request.POST.items():
            if (name == 'fecha'):
                 year = int(value)
        tramites=PlanillaDeInspeccion.objects.values('tramite_id').distinct().order_by('tramite_id').annotate(Max('fecha'))
        for i in item:
            totalI=0
            subtotal=0
            for mes in range(12):
                m = mes + 1
                diaFinal = monthrange(year, m)
                subtotal=0
                for t in tramites:
                    try:
                        resultado=PlanillaDeInspeccion.objects.filter(fecha=t['fecha__max'],tramite_id=t['tramite_id'], fecha__range=(
                        datetime.date(year, m, 01), datetime.date(year, m, diaFinal[1])), detalles__nombre=i.nombre, detalles__id=i.id)
                        if resultado:
                           subtotal+=1
                    except:
                        pass
                totalI+=subtotal
                cant.append(subtotal)
            datos.append(tuple(cant))
            nombre=str(i.nombre)
            series.append(nombre)
            totalItems=PlanillaDeInspeccion.objects.filter(detalles__id=i.id, fecha__range=(datetime.date(year,01, 01), datetime.date(year, 12, 31))).values('tramite_id').distinct().order_by('tramite_id').annotate(Max('fecha')).count()
            # if (totalItems==0):
            #      porcentaje=0
            #      porcentaje1=0
            # else:
            #      porcentaje1= (totalI/float(totalItems))*100
            #      porcentaje = "{0:.2f}".format(porcentaje1)
            aux = {'nombre':nombre,'categoria':i.categoria_inspeccion.nombre,'item':i.item_inspeccion.nombre, 'cantidad': totalItems}
            lista.append(aux)
            cant = []
        titulo = "Items de inspeccion por mes"
        if len(datos) > 0:
            grafico = grafico_de_barras_v(datos, nombres, titulo,series)
            imagen = base64.b64encode(grafico.asString("png"))
            contexto = {"grafico": imagen,"items":item,"lista":lista}
            return render(request, 'persona/director/listado_item_inspeccion.html', contexto)

        else:
            return render(request, 'persona/director/listado_item_inspeccion.html')
    else:
        categorias=DetalleDeItemInspeccion.objects.select_related().values('categoria_inspeccion_id','categoria_inspeccion__nombre','item_inspeccion_id','item_inspeccion__nombre','activo').order_by('categoria_inspeccion_id','item_inspeccion_id').distinct()
        items=DetalleDeItemInspeccion.objects.select_related().values('id','nombre','categoria_inspeccion_id','item_inspeccion_id','item_inspeccion__nombre','activo').order_by('categoria_inspeccion_id','item_inspeccion_id').all()
        return render(request, 'persona/director/seleccionar_item_fecha.html',{"items":items,"categorias":categorias})

def devolverNombreMes(m):
    nombres = [["Enero",1], ["Febrero",2], ["Marzo",3], ["Abril",4], ["Mayo",5],
               ["Junio",6], ["Julio",7], ["Agosto",8], ["Septiembre",9], ["Octubre",10],
                   ["Noviembre",11], ["Diciembre",12]]
    mes = ""
    for name,value in nombres:
        if value == m:
            mes = name
            return mes

def tramites_iniciados_finalizados(request):
        datos = []
        iniciados = []
        finalizados = []
        series = []
        lista = []
        lista1 = []
        lista2 = []
        nombres = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre",
                   "Noviembre", "Diciembre"]
        if "Guardar" in request.POST:
            for name, value in request.POST.items():
                if name.startswith('item'):
                    year = int(value)
                if name.startswith('obra'):
                    tipoObra=int(value)
            totalIniciados = 0
            totalFinalizados = 0
            for mes in range(12):
                m = mes + 1
                nombreMes = ""
                diaFinal = monthrange(year, m)
                totalI = Estado.objects.filter(tramite__tipo_obra=tipoObra,timestamp__range=(datetime.date(year, m, 01), datetime.date(year, m, diaFinal[1])),
                    tipo=(1)).count()
                totalF = Estado.objects.filter(tramite__tipo_obra=tipoObra, timestamp__range=(datetime.date(year, m, 01), datetime.date(year, m, diaFinal[1])), tipo=(9)).count()
                iniciados.append(totalI)
                totalIniciados=totalI+totalIniciados
                totalFinalizados=totalF+totalFinalizados
                finalizados.append(totalF)
                nombreMes=devolverNombreMes(m)
                aux = {"id": m,"mes": nombreMes, "iniciado":totalI, "finalizado":totalF}
                lista1.append(aux)
            i = tuple(iniciados)
            f = tuple(finalizados)
            datos.append(i)
            datos.append(f)
            iniciales = Estado.objects.filter(tipo=1)
            finalizados = Estado.objects.filter(tipo=9).count()
            inicial=0
            for i in iniciales:
                if i.previo() is None:
                  inicial =inicial+1
            finales = Estado.objects.filter(tipo=9).count()
            porcentajeIni=(totalIniciados/float(inicial))*100
            porcentajeI="{0:.2f}".format(porcentajeIni)
            porcentajeFini=(totalFinalizados/float(inicial))*100
            porcentajeF = "{0:.2f}".format(porcentajeFini)
            lista.append(["iniciados", porcentajeI])
            lista.append(["finalizados", porcentajeF])
            series = ("iniciados", "finalizados")
            titulo = "Tramites iniciados y finalidos por mes"
            if len(datos) > 0:
                grafico = grafico_de_barras_v(datos, nombres, titulo, series)
                imagen = base64.b64encode(grafico.asString("png"))
                contexto = {"grafico": imagen, "lista": lista, "lista1":lista1}  # "tipos_obras": list}
            return render(request, 'persona/director/listado_tramites_iniciados_finalizados.html', contexto)
        else:
            tipos_obras=TipoObra.objects.filter(activo=1)
            return render(request, 'persona/director/seleccionar_fecha.html', {"tipos_obras":tipos_obras})


def recaudacion_tipo_obra_tipo_pago(request):
        datos = []
        lista = []
        totalIniciados = 0
        totalFinalizados = 0
        tramites = []
        totalAnual = 0
        pagos = []
        cuotas = []
        totales = []
        tipoP = ""
        nombreP=""
        tipoO = ""
        nombres = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre",
                   "Noviembre", "Diciembre"]
        if "Guardar" in request.POST:
            for name, value in request.POST.items():
                if name.startswith('item'):
                    year = int(value)
                if name.startswith('obra'):
                    tipoObra = int(value)
                if name.startswith('tipoPago'):
                    tipoPago = int(value)
            tramites=Tramite.objects.filter(tipo_obra=tipoObra).values_list('pago_id', flat="True").distinct() #filtrar por tipo
            for mes in range(12):
                m = mes + 1
                diaFinal = monthrange(year, m)
                if tipoPago>-1:
                    cuotas = Cuota.objects.filter(fechaPago__range=(datetime.date(year, m, 01), datetime.date(year, m, diaFinal[1])),pago_id__in=tramites, tipoPago_id=tipoPago)
                else:
                    cuotas = Cuota.objects.filter(fechaPago__range=(datetime.date(year, m, 01), datetime.date(year, m, diaFinal[1])),pago_id__in=tramites)
                total=0
                for c in cuotas:
                    total=float(c.monto)+total
                totales.append(total)
                totalAnual=total+totalAnual
                aux={"id":m,"mes":nombres[mes],"total":total}
                lista.append(aux)
            datos.append(totales)
            series = ["Recaudacion"]
            tipoO=TipoObra.objects.get(id=tipoObra)
            if tipoPago > -1:
                tipoP = Tipo_Pago.objects.get(id=tipoPago)
                nombreP=tipoP.nombre
            if len(datos) > 0:
                grafico = grafico_de_barras_v(datos, nombres, "", series)
                imagen = base64.b64encode(grafico.asString("png"))
                contexto = {"grafico": imagen,"lista":lista, "totalAnual":totalAnual,"tipoP":nombreP,"tipoO":tipoO.nombre}
            return render(request, 'persona/director/recaudacion_tipo_obra_tipo_pago.html', contexto)
        else:
            tipos_obras = TipoObra.objects.filter(activo=1)
            tipo_pago = Tipo_Pago.objects.filter(activo=1)
        return render(request, 'persona/director/seleccionar_datos.html', {"tipos_obras": tipos_obras,"tipo_pago":tipo_pago})


def ver_sectores_con_mas_obras(request):
    tramites = Tramite.objects.all()
    sectores = []
    list = []
    for t in tramites:
        if not t.sector in sectores:
            sectores.append(t.sector)

    for s in sectores:
        list.append([s, 0])

    sectores = list
    list_sectores = []

    for name, value in sectores:
        v = 0
        for t in tramites:
            if t.sector == name:
                v += 1
        list_sectores.append([name, v])
    contexto = {
        "sectores": list_sectores,
        "nombres": list
    }
    return render(request,'persona/director/ver_sectores_con_mas_obras.html',contexto)


def seleccionar_tipoObra_sector(request):
    datos = []
    iniciados = []
    series = []
    lista = []
    nombres = ["Sectores"]
    tramites = Tramite.objects.all()
    tiposObras = TipoObra.objects.filter(activo=1)
    if "Guardar" in request.POST:
        for name, value in request.POST.items():
            if name.startswith('obra'):
                tipoObra = int(value)
        for to in tiposObras:
            if to.id == tipoObra:
                tipo_obra = to.nombre
        sectores = []
        list = []
        tramite = Tramite.objects.filter(tipo_obra_id=tipoObra)
        for t in tramite:
                if not t.sector in sectores:
                    sectores.append(t.sector)
        for s in sectores:
               list.append([s, 0])
        sectores = list
        list_sectores = []
        listaSectores = []
        for name, value in sectores:
                v = 0
                for t in tramites:
                    if t.sector == name:
                        v += 1
                list_sectores.append([v,name])
             #   listaSectores.append(v)
        list_sectores.sort(reverse=True)
        for l in list_sectores[0:30]:
            datos.append([l[0]])
            nombre="Sector "+ str(l[1])
            nombres.append(nombre)
            series.append(nombre)
        titulo = "Sectores con mas obras segun obra seleccionada"
        if len(datos) > 0:
                grafico = grafico_de_barras_v(datos, nombres, titulo,series)
                imagen = base64.b64encode(grafico.asString("png"))
                contexto = {"grafico": imagen, "lista": list_sectores[0:30], "tipo_obra":tipo_obra}
        return render(request, 'persona/director/sectores_con_mas_obras.html', contexto)
        #except:
        #   contexto={"respuesta":"No hay datos para mostrar"}
         #   return render(request, 'persona/director/sectores_con_mas_obras.html', contexto)

    else:
        tramObras = Tramite.objects.all()
        tramites = Tramite.objects.all().values_list('tipo_obra_id', flat="True").distinct()
        tipos_obras = TipoObra.objects.filter(id__in=tramites)
        return render(request, 'persona/director/seleccionar_tipoObra_sector.html', {"tipos_obras": tipos_obras})

def ver_listado_usuarios(request):
    grupossistema = Group.objects.all()
    personas = Persona.objects.all()
    profesionales = Profesional.objects.all()
    propietarios = Propietario.objects.all()
    listados = []
    for g in grupossistema:
        for p in personas:
            for pr in propietarios:
                for pf in profesionales:
                    if p.usuario_id == g.id or pr.id == g.id or pf.id == g.id:
                        if p not in listados:
                            listados.append(p)

    contexto = {'listados':listados, 'grupossistema':grupossistema}
    return render(request, 'persona/director/listado_de_usuarios_segun_grupo.html', contexto)

def tiempo_aprobacion_visados(request):
    datos = []
    visados = []
    auxiliar=[]
    aprobados = []
    tramitesAgendados=[]
    tramitesAprobados=[]
    agendados=[]
    tramA = []
    tramAg = []
    cant = 0
    lista = []
    lista2=[]
    tramites = []
    tramitesV = []
    meses = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
    listaAp = []
    nombres = ["1 mes", "2 meses", "3 meses", "4 meses", "5 meses", "6 meses", "7 meses", "8 meses", "9 meses", "10 meses", "11 meses", "12 meses"]
    if "Guardar" in request.POST:
        for name, value in request.POST.items():
            if name.startswith('fecha'):
                year = int(value)
        tramites=Estado.objects.filter(timestamp__range=(datetime.date(year, 01, 01),(datetime.date(year, 12, 31))),tipo=1).values_list('tramite', flat="True").distinct()
        tramitesV=Estado.objects.filter(timestamp__range=(datetime.date(year, 01, 01),(datetime.date(year, 12, 31))),tipo=5,tramite_id__in=tramites).values_list('tramite', flat="True").distinct() #todos los tramites que terminaron de ser visados (anuales)
        tramitesAprobados=Estado.objects.filter(timestamp__range=(datetime.date(year, 01, 01),(datetime.date(year, 12, 31))),tipo=3,tramite_id__in=tramites).values_list('tramite', flat="True").distinct() #todos los tramites que terminaron de ser visados (anuales)
        tramitesAgendados=Estado.objects.filter(timestamp__range=(datetime.date(year, 01, 01),(datetime.date(year, 12, 31))),tipo=5, tramite_id__in=tramites).values_list('tramite', flat="True").distinct() #todos los tramites que terminaron de ser visados (anuales)
        listaAg=[]
        for mes in range(12):
            m = mes + 1
            cant = 0
            agendados=[]
            diaFinal = monthrange(year, m)
            tEstadoAprobado=Estado.objects.filter(timestamp__range=(datetime.date(year, m, 01), datetime.date(year, m, diaFinal[1])), tipo=(3), tramite_id__in=tramites).distinct()#.count()
            tramA=[t.tramite_id for t in tEstadoAprobado]
            aprobados=list(set(tramA))
            tEstadoAgendado = Estado.objects.filter(timestamp__range=(datetime.date(year, m, 01), datetime.date(year, m, diaFinal[1])),tipo=(5), tramite_id__in=tramites).distinct()  # .count()
            for t in tEstadoAgendado:
                if t.tramite_id not in auxiliar:
                    auxiliar.append(t.tramite_id)
                    agendados.append(t.tramite_id)
            a={"mes":mes,"aprobados":aprobados}
            b={"mes":mes,"agendados":agendados}
            listaAp.append(a)
            listaAg.append(b)
        for i in listaAp:
            for j in i["aprobados"]:
                for s in listaAg:
                    for d in s["agendados"]:
                        diferencia=s["mes"]-i["mes"]
                        if j==d:
                             if diferencia >=0:
                                 try:
                                     cant=meses[diferencia]+1
                                     meses[diferencia]=cant
                                 except:
                                     pass
        if tramites.count() == 0:
            messages.add_message(request, messages.SUCCESS, 'No hay visados aprobados en el año seleccionado')
            return render(request, 'persona/director/seleccionar_fecha_visados_aprobados.html')
        else:
            meses=[cant/float(tramites.count()) for cant in meses]
            planillas= PlanillaDeVisado.objects.filter(fecha__range=(datetime.date(year, 01, 01),(datetime.date(year, 12, 31))), tramite_id__in=tramites)
            for t in tramitesV:
                aux=filter(lambda p: t==p.tramite_id, planillas)
                if (len(aux))!=0:
                    lista.append([t,len(aux)])
            datos.append(meses)
            porcentajeVisadosAprobados = ((tramitesAprobados.count() / float(tramites.count())) * 100)
            porcentajeApr = "{0:.2f}".format(porcentajeVisadosAprobados)
            lista2.append(["Aprobados", porcentajeApr])
            porcentajeVisadosFinalizados = ((tramitesAgendados.count() / float(tramites.count())) * 100)
            porcentajeFin = "{0:.2f}".format(porcentajeVisadosFinalizados)
            lista2.append(["Finalizados", porcentajeFin])
            titulo = "Promedio de duracion (en meses) de inicio y finalizacion de visados"
            if len(datos) > 0:
                  grafico = grafico_de_barras_v(datos, nombres, titulo,["promedio"])
                  imagen = base64.b64encode(grafico.asString("png"))
                  contexto = {"grafico": imagen, "lista": lista,"Promedio":lista2}
            return render(request, 'persona/director/tiempo_aprobacion_visados.html',contexto)

    else:
        return render(request, 'persona/director/seleccionar_fecha_visados_aprobados.html')

class ReporteTramitesDirectorExcel(TemplateView):

    def get(self, request, *args, **kwargs):
        tramites = Tramite.objects.all()
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'REPORTE DE TRAMITES'
        ws.merge_cells('B1:G1')
        ws['B2'] = 'NRO'
        ws['C2'] = 'PROPIETARIO'
        ws['D2'] = 'PROFESIONAL'
        ws['E2'] = 'ESTADO'
        ws['F2'] = 'MEDIDAS'
        ws['G2'] = 'TIPO DE OBRA'
        cont = 3
        for tramite in tramites:
            ws.cell(row=cont, column=2).value = tramite.id
            ws.cell(row=cont, column=3).value = str(tramite.propietario)
            ws.cell(row=cont, column=4).value = str(tramite.profesional)
            ws.cell(row=cont, column=5).value = str(tramite.estado())
            ws.cell(row=cont, column=6).value = tramite.medidas
            ws.cell(row=cont, column=7).value = str(tramite.tipo_obra)
            cont = cont + 1
        nombre_archivo = "ReporteTramites.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response

class ReporteTramitesDirectorPdf(View):

    def get(self, request, *args, **kwargs):
        filename = "Informe de tramites " + datetime.datetime.now().strftime("%d/%m/%Y") + ".pdf"
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="%s"' % filename
        doc = SimpleDocTemplate(
            response,
            pagesize=letter,
            rightMargin=72,
            leftMargin=72,
            topMargin=15,
            bottomMargin=28,
        )
        Story = []
        styles = getSampleStyleSheet()
        styles.add(ParagraphStyle(name='Usuario', alignment=TA_RIGHT, fontName='Helvetica', fontSize=8))
        styles.add(ParagraphStyle(name='Titulo', alignment=TA_RIGHT, fontName='Helvetica', fontSize=18))
        styles.add(ParagraphStyle(name='Subtitulo', alignment=TA_RIGHT, fontName='Helvetica', fontSize=12))
        usuario = 'Usuario: ' + request.user.username + ' -  Fecha:' + str(datetime.date.today())
        Story.append(Paragraph(usuario, styles["Usuario"]))
        Story.append(Spacer(0, cm * 0.15))
        im0 = Image(settings.MEDIA_ROOT + '/imagenes/espacioPDF.png', width=490, height=3)
        im0.hAlign = 'CENTER'
        Story.append(im0)
        titulo = 'SISTEMA OBRAS PARTICULARES'
        Story.append(Paragraph(titulo, styles["Titulo"]))
        Story.append(Spacer(0, cm * 0.20))
        subtitulo = 'Reporte de tramites'
        Story.append(Paragraph(subtitulo, styles["Subtitulo"]))
        Story.append(Spacer(0, cm * 0.15))
        Story.append(im0)
        Story.append(Spacer(0, cm * 0.5))
        encabezados = ('NRO', 'PROPIETARIO', 'PROFESIONAL', 'ESTADO', 'MEDIDAS', 'TIPO DE OBRA')
        detalles = [
            (tramite.id, tramite.propietario, tramite.profesional, tramite.estado(), tramite.medidas, tramite.tipo_obra)
            for tramite in
            Tramite.objects.all()]
        detalle_orden = Table([encabezados] + detalles, colWidths=[1 * cm, 3 * cm, 4 * cm, 4 * cm, 2 * cm, 3 * cm])
        detalle_orden.setStyle(TableStyle(
            [
                ('ALIGN', (0, 0), (0, 0), 'CENTER'),
                ('GRID', (0, 0), (-1, -1), 1, colors.gray),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('LINEBELOW', (0, 0), (-1, 0), 2, colors.darkblue),
                ('BACKGROUND', (0, 0), (-1, 0), colors.dodgerblue),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ]
        ))
        detalle_orden.hAlign = 'CENTER'
        Story.append(detalle_orden)
        doc.build(Story)
        return response

class ReporteInspeccionesDirectorExcel(TemplateView):
    def get(self, request, *args, **kwargs):
        wb = Workbook()
        ws = wb.active
        ws.merge_cells('B1:Q1')
        ws.merge_cells('C2:F2')
        ws.merge_cells('G2:K2')
        ws.merge_cells('L2:Q2')
        ws.merge_cells('C3:D3')
        ws.merge_cells('E3:F3')
        ws.merge_cells('H3:I3')
        ws.merge_cells('J3:K3')
        ws.merge_cells('L3:M3')
        ws.merge_cells('N3:O3')
        ws.merge_cells('P3:Q3')
        ws['B1'] = 'REPORTE DE INSPECCIONES ANUALES'
        ws['C2'] = 'TRAMITE'
        ws['G2'] = 'PROPIETARIO'
        ws['L2'] = 'PROFESIONAL'
        ws['B3'] = 'NRO'
        ws['C3'] = 'DOMICILIO'
        ws['E3'] = 'FECHA'
        ws['G3'] = 'DNI'
        ws['H3'] = 'NOMBRE'
        ws['J3'] = 'APELLIDO'
        ws['L3'] = 'MATRICULA'
        ws['N3'] = 'NOMBRE'
        ws['P3'] = 'APELLIDO'
        cont = 4
        year = date.today()
        tramitesEstado = Estado.objects.select_related().values('timestamp', 'tramite__id','tramite__domicilio','tramite__propietario__persona__dni','tramite__propietario__persona__nombre','tramite__propietario__persona__apellido','tramite__profesional__matricula','tramite__profesional__persona__nombre','tramite__profesional__persona__apellido').filter(
            timestamp__range=(datetime.date(year.year, 01, 01), datetime.date(year.year, 12, 12)), tipo=(6)).exclude(id__isnull=True)
        for t in tramitesEstado:
            ws.cell(row=cont, column=2).value = t['tramite__id']
            ws.cell(row=cont, column=3).value = t['tramite__domicilio']
            ws.cell(row=cont, column=5).value = str(t['timestamp'])
            ws.cell(row=cont, column=7).value = t['tramite__propietario__persona__dni']
            ws.cell(row=cont, column=8).value = t['tramite__propietario__persona__nombre']
            ws.cell(row=cont, column=10).value = t['tramite__propietario__persona__apellido']
            ws.cell(row=cont, column=12).value = t['tramite__profesional__matricula']
            ws.cell(row=cont, column=14).value = t['tramite__profesional__persona__nombre']
            ws.cell(row=cont, column=16).value = t['tramite__profesional__persona__apellido']
            cont = cont + 1
        nombre_archivo = "ReporteInspecciones.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response

class ReporteInspeccionesDirectorPdf(View):
    def get(self, request, *args, **kwargs):
        filename = "Informe de Inspecciones " + datetime.datetime.now().strftime("%d/%m/%Y") + ".pdf"
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="%s"' % filename
        doc = SimpleDocTemplate(
            response,
            pagesize=letter,
            rightMargin=72,
            leftMargin=72,
            topMargin=15,
            bottomMargin=28,
        )
        Story = []
        styles = getSampleStyleSheet()
        styles.add(ParagraphStyle(name='Usuario', alignment=TA_RIGHT, fontName='Helvetica', fontSize=8))
        styles.add(ParagraphStyle(name='Titulo', alignment=TA_RIGHT, fontName='Helvetica', fontSize=18))
        styles.add(ParagraphStyle(name='Subtitulo', alignment=TA_RIGHT, fontName='Helvetica', fontSize=12))
        usuario = 'Usuario: ' + request.user.username + ' -  Fecha:' + str(datetime.date.today())
        Story.append(Paragraph(usuario, styles["Usuario"]))
        Story.append(Spacer(0, cm * 0.15))
        im0 = Image(settings.MEDIA_ROOT + '/imagenes/espacioPDF.png', width=490, height=3)
        im0.hAlign = 'CENTER'
        Story.append(im0)
        titulo = 'SISTEMA OBRAS PARTICULARES'
        Story.append(Paragraph(titulo, styles["Titulo"]))
        Story.append(Spacer(0, cm * 0.20))
        subtitulo = 'Reporte de Inspecciones'
        Story.append(Paragraph(subtitulo, styles["Subtitulo"]))
        Story.append(Spacer(0, cm * 0.15))
        Story.append(im0)
        Story.append(Spacer(0, cm * 0.5))
        encabezados = ('NRO', 'DOMICILIO', 'FECHA', 'DNI', 'NOMBRE', 'APELLIDO','MP','NOMBRE','APELLIDO')
        e_titulos = ('TRAMITE',  'PROPIETARIO','PROFESIONAL')
        year = date.today()
        tramitesEstado = Estado.objects.select_related().values('timestamp', 'tramite__id', 'tramite__domicilio',
                                                                'tramite__propietario__persona__dni',
                                                                'tramite__propietario__persona__nombre',
                                                                'tramite__propietario__persona__apellido',
                                                                'tramite__profesional__matricula',
                                                                'tramite__profesional__persona__nombre',
                                                                'tramite__profesional__persona__apellido').filter(
            timestamp__range=(datetime.date(year.year, 01, 01), datetime.date(year.year, 12, 12)), tipo=(6)).exclude(id__isnull=True)
        detalles = [
            (tramite['tramite__id'], tramite['tramite__domicilio'], str(tramite['timestamp']),tramite['tramite__propietario__persona__dni'],
            tramite['tramite__propietario__persona__nombre'], tramite['tramite__propietario__persona__apellido'], tramite['tramite__profesional__matricula'],
            tramite['tramite__profesional__persona__nombre'],tramite['tramite__profesional__persona__apellido'])
            for tramite in tramitesEstado ]
        detalle=Table([e_titulos], colWidths=[8.5* cm,6.6*cm,6.1*cm] )
        detalle.setStyle(TableStyle(
            [
                ('ALIGN', (0, 0), (0, 0), 'LEFT'),
                ('GRID', (0, 0), (-1, -1), 1, colors.gray),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('LINEBELOW', (0, 0), (-1, 0), 2, colors.darkblue),
                ('BACKGROUND', (0, 0), (-1, 0), colors.dodgerblue),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ]
        ))
        detalle_orden = Table([encabezados]+ detalles, colWidths=[1 * cm, 2.5 * cm, 5 * cm, 1.6 * cm, 2.5 * cm, 2.5 * cm,1.5*cm,2.3*cm,2.3*cm])
        detalle_orden.setStyle(TableStyle(
            [
                ('ALIGN', (0, 0), (0, 0), 'CENTER'),
                ('GRID', (0, 0), (-1, -1), 1, colors.gray),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('LINEBELOW', (0, 0), (-1, 0), 2, colors.darkblue),
                ('BACKGROUND', (0, 0), (-1, 0), colors.dodgerblue),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ]
        ))
        detalle_orden.hAlign = 'CENTER'
        Story.append(detalle)
        Story.append(detalle_orden)
        doc.build(Story)
        return response
#-------------------------------------------------------------------------------------------------------------------
#No se de donde son estos------------------------------------------------------------------

def tramite_visados_list(request):
    tramites = Tramite.objects.en_estado(Visado) #cambiar a visados cuando etengas tramites visaddos
    contexto = {'tramites': tramites}
    return contexto

def mostrar_popup_datos_agendar(request,pk_tramite):
    pass

def alta_persona(request):
    if request.method == "POST":
        form = FormularioPersona(request.POST)
        if form.is_valid():
            persona = form.save()
            persona.save()
    else:
        form = FormularioPersona()
    return render(request, 'persona/alta/alta_persona.html', {'form': form})

#------------------------------------------------------------------------------------------------------------------
#cajero ---------------------------------------------------------------------------------------------------------
from tipos import forms
from tipos import models

@login_required(login_url="login")
@grupo_requerido('cajero')
def mostrar_cajero(request):
    contexto = {
        "ctxtramites_para_financiar": listado_tramites_para_financiar(request),
        "ctxcuotas":listado_tramites_a_pagar(request),
        "ctxlistado":listado_tramites(request),
    }
    return render(request, 'persona/cajero/cajero.html', contexto)

def listado_de_comprobantes(request, pk_tramite):
        tramite = get_object_or_404(Tramite, pk=pk_tramite)
        pago = tramite.pago
        canceladas = []
        cuotas = Cuota.objects.en_estado(Cancelada)
        for cuota in cuotas:
            if cuota.pago == pago:
                canceladas.append(cuota)
        if canceladas is None:
            messages.add_message(request, messages.WARNING, 'No hay pagos registrados para el tramite seleccionado.')
        return render(request, 'persona/cajero/listado_de_comprobantes.html', {'cuotas': canceladas,'tramite':tramite})


def listado_tramites_para_financiar(request):
    tramite = Tramite.objects.en_estado([Visado,Agendado,ConInspeccion,Inspeccionado,FinalObraSolicitado])
    listado=[]
    for tramites in tramite:
        if tramites.pago is None:
           listado.append(tramites)
    contexto = {'tramites':listado}
    return contexto

def elegir_financiacion(request,pk_tramite):
    tramite = get_object_or_404(Tramite, pk=pk_tramite)
    if request.method == "POST":
        if "Guardar" in request.POST:
            pago = Pago()
            contador = 31
            fms = "%A"
            for name, value in request.POST.items():
                if name.startswith('cantidadCuotas'):
                    pago.cantidadCuotas=value
            total = tramite.monto_a_pagar/int(pago.cantidadCuotas)
            pago.save()
            for i in range(1, int(pago.cantidadCuotas)+1):
                cuota = Cuota(monto=total, numeroCuota=i, pago=pago)
                cuota.fechaVencimiento=date.today() + timedelta(days=contador)
                dia=cuota.fechaVencimiento.strftime(fms)
                if dia=="Sunday":
                    cuota.fechaVencimiento==date.today() + timedelta(days=contador+1)
                else:
                    if dia=="Saturday":
                        cuota.fechaVencimiento = date.today() + timedelta(days=contador +2)
                contador=contador+31
                cuota.save()
                cuota.hacer("Cancelacion")
            messages.add_message(request, messages.SUCCESS, 'Se ha guardado la cantidad de cuotas para el pago seleccionada')
            tramite.pago = pago
            tramite.save()
        return redirect('cajero')
    return render(request, 'persona/cajero/elegir_financiacion.html',{'tramite': tramite})

def registrar_pago(request,pk_tramite):
    tramite = get_object_or_404(Tramite, pk=pk_tramite)
    if request.method == "POST":
        form = FormularioPago(request.POST)
        if form.is_valid():
            pago = form.save(commit=False)
            contador=31
            fms = "%A"
            if tramite.pago is None:
                pago.save()
                tramite.pago = pago
                tramite.save()
                total=tramite.monto_a_pagar/pago.cantidadCuotas
                for i in range(1, pago.cantidadCuotas+1):
                    cuota = Cuota(monto=total,numeroCuota=i,pago=pago)
                    cuota.fechaVencimiento=date.today() + timedelta(days=contador)
                    dia=cuota.fechaVencimiento.strftime(fms)
                    if (dia=="Sunday"):
                        cuota.fechaVencimiento=date.today() + timedelta(days=contador+1)
                    else:
                        if (dia=="Saturday"):
                            cuota.fechaVencimiento = date.today() + timedelta(days=contador + 2)
                    contador=contador+31
                    cuota.save()
                    cuota.hacer("Cancelacion")
            else:
                messages.add_message(request, messages.ERROR, 'El tramite ya tiene un pago registrado.')
    else:
        form = FormularioPago()
    return form

def listado_tramites_a_pagar(request):
    objetos=Tramite.objects.all()
    tramites=[]
    for tramite in objetos:
        if ((tramite.pago is not None) and (tramite.esta_pagado()==False)):
            tramites.append(tramite)
    contexto={'tramites':tramites}
    return contexto

def listado_cuotas(request):
    cuotas=Cuota.objects.en_estado(Cancelacion)
    contexto= {'cuotas':cuotas}
    return contexto

def elegir_tramite(request, pk_tramite):
    tramite=get_object_or_404(Tramite,pk=pk_tramite)
    pago=tramite.pago
    cuotas=[]
    c=None
    objetos=Cuota.objects.en_estado(Cancelacion)
    for cuota in objetos:
        if cuota.fechaPago is None and cuota.pago==pago:
            c=cuota
            break;
    return render(request, 'persona/cajero/registrar_cuota.html', {'cuotas':c, 'tramite':tramite})

def pagarCuota(cuota):
    cuota.guardar_fecha()
    cuota.save()
    cuota.hacer("cancelacion")
    pago = cuota.pago
    tramite = get_object_or_404(Tramite, pago=pago)
    tramite.calcular_monto_pagado(cuota.monto)
    tramite.save()
    return tramite

def pagar_cuota(request,pk_cuota):
    cuota = get_object_or_404(Cuota, pk=pk_cuota)
    tiposPagos = Tipo_Pago.objects.all()
    return render(request, 'persona/cajero/pagar_cuota.html', {'cuota':cuota, 'tiposPagos':tiposPagos})

def pagar1(request, pk_cuota):
    cuota = get_object_or_404(Cuota, pk=pk_cuota)
    datosPagos = Tipo_Pago.objects.all()
    tipoPago = 0
    if "Guardar" in request.POST:
        for n in datosPagos:
            if request.POST['tipoPago'] == n.nombre:
                tipoPago = n.id
    tp = Tipo_Pago.objects.get(id=tipoPago)
    cuota.guardar_fecha()
    pago = cuota.pago
    cuota.tipoPago = tp
    cuota.hacer("cancelacion")
    cuota.save()
    tramite = get_object_or_404(Tramite, pago=pago)
    tramite.calcular_monto_pagado(cuota.monto)
    tramite.save()
    messages.add_message(request, messages.SUCCESS, 'Pago Registrado.')
    contexto = {'tramite': tramite, 'pago': pago, 'cuota': cuota}
    return render(request, 'persona/cajero/registrar_pago_tramite.html', contexto)

def pagar(request):
    c = []
    for name, value in request.POST.items():
        if name.startswith('cuota'):
            pk = name.split('-')[1]
            c.append(pk)
    p = []
    cuotas = []
    for cs in c:
        cuota = get_object_or_404(Cuota, pk=cs)
        cuotas.append(cuota)
        tramite=pagarCuota(cuota)
    pago=cuota.pago
    messages.add_message(request, messages.SUCCESS, 'Pago Registrado.')
    contexto = {'tramite': tramite, 'pago': pago, 'cuota': cuotas}
    return render ( request,'persona/cajero/registrar_pago_tramite.html', contexto)

def elegir_cuota(request,pk_cuota):
    cuota=get_object_or_404(Cuota,pk=pk_cuota)
    cuota.guardar_fecha()
    cuota.save()
    cuota.hacer("cancelacion")
    pago = cuota.pago
    tramite = get_object_or_404(Tramite, pago=pago)
    tramite.calcular_monto_pagado(cuota.monto)
    tramite.save()
    messages.add_message(request, messages.SUCCESS, 'Pago Registrado.')
    return render(request, 'persona/cajero/actualizar_cuota.html',{'cuota':cuota})

def comprobante_pago_cuota(request,pk_cuota):
    cuota = get_object_or_404(Cuota, pk=pk_cuota)
    pago = cuota.pago
    tramite = get_object_or_404(Tramite, pago=pago)
    return render(request, 'persona/cajero/comprobante.html',{'cuota': cuota, 'pago':pago,'tramite':tramite})

def registrar_el_pago_tramite(request):
    c = []
    for name, value in request.POST.items():
        if name.startswith('cuota'):
            pk = name.split('-')[1]
            c.append(pk)
    p = []
    cuotas = []
    for cs in c:
        cuota = get_object_or_404(Cuota, pk=cs)
        cuotas.append(cuota)
    pago = cuota.pago
    tramite = get_object_or_404(Tramite, pago=pago)
    contexto = {'tramite': tramite,'pago':pago, 'cuota': cuotas}
    return render ( request,'persona/cajero/registrar_pago_tramite.html', contexto)

def listado_tramites(request):
    objetos=Tramite.objects.all()
    tramites=[]
    for tramite in objetos:
        if (tramite.pago is not None and tramite.monto_pagado>0):
            tramites.append(tramite)
    contexto={'tramites':tramites}
    return contexto

def listado_comprobantes(request,pk_tramite):
    tramite=get_object_or_404(Tramite,pk=pk_tramite)
    pago=tramite.pago
    canceladas=[]
    cuotas=Cuota.objects.en_estado(Cancelada)
    for cuota in cuotas:
        if cuota.pago==pago:
            canceladas.append(cuota)
    if canceladas is None:
        messages.add_message(request, messages.WARNING, 'No hay pagos registrados para el tramite seleccionado.')
    return render (request, 'persona/cajero/factura_parcial.html', {'cuotas':canceladas,'tramite':tramite,'pago':pago})


#------------------------------------------------------------------------------------------------------------------
#movil ---------------------------------------------------------------------------------------------------------
#
# def movil_login(request):
#     return render(request, 'movil/templates/login.html', contexto)
def es_inspector(usuario):
    return usuario.groups.filter(name='inspector' or 'jefeinspector').exists()

#@grupo_requerido('inspector' or 'jefeinspector')
@user_passes_test(es_inspector)

def mostrar_inspector_movil(request):
    usuario = request.user
    if (request.user_agent.is_mobile):
        contexto = {
            "ctxlistado_inspector": listado_inspector_movil(request)
        }
    else:
        return redirect('inspector')
    return render(request, 'persona/movil/inspector_movil.html',contexto)

def mostrar_inspector_movil_jefe(request):
    usuario = request.user
    if (request.user_agent.is_mobile):
        contexto = {
            "ctxlistado_inspector": listado_inspecciones(request)#listado_inspector_movil(request)
        }
        usuario = request.user
        if (usuario.groups.filter(name='inspector').exists()):
                contexto = {
                    "ctxlistado_inspector": listado_inspector_movil(request),
                    "ctxlistadomensual_inspector": listado_inspecciones_mensuales(request),
                }
        else:
                if (usuario.groups.filter(name='jefeinspector').exists()):
                    contexto = {
                        "ctxlistado_inspector": listado_inspecciones(request),
                        "ctxlistadomensual_inspector":listado_inspecciones_mensuales_jefe_inspector(request),
                    }
    else:
        return redirect('jefeinspector')
    return render(request, 'persona/movil/inspector_movil.html',contexto)

def movil_inspector(request):
    return render(request, 'persona/movil/inspector_movil.html')

def frente_o_fachada(request):
    return render(request,'persona/movil/frente_o_fachada.html')    

def paredes(request):    
    return render(request,'persona/movil/paredes.html')    

def cocinas(request):    
    return render(request,'persona/movil/cocinas.html')    

def techos(request):    
    return render(request,'persona/movil/techos.html')                

def listado_inspector_movil(request):
    usuario = request.user
    estados = Estado.objects.all()
    tipo = 5 #Agendados
    argumentos = [Visado, ConInspeccion]
    tramites_del_inspector = Tramite.objects.en_estado(Agendado)
    tramites = filter(lambda t: t.estado().usuario == usuario and t.estado().rol==2, tramites_del_inspector)
    contexto={'tramites':tramites}
    return contexto

def listado_inspecciones_mensuales(request):
    year=datetime.date.today().year
    mes=datetime.date.today().month
    dia=datetime.date.today().day
    diaFinal=monthrange(year, mes)
    usuario = request.user
    tramites_del_inspector = Tramite.objects.en_estado(Agendado)
    tramites = filter(lambda t: t.estado().usuario == usuario and t.estado().rol==2 and t.estado().fecha.date()<=datetime.date(year,mes,diaFinal[1]) and t.estado().fecha.date()>=datetime.date(year,mes,dia), tramites_del_inspector)
    contexto={'tramites':tramites}
    return contexto

def planilla_inspeccion_movil(request,pk_tramite):
    tramite = get_object_or_404(Tramite, pk=pk_tramite)
    movil=es_movil(request)
    planilla=PlanillaDeInspeccion.objects.select_related().filter(tramite_id=tramite).last() #last una sola planilla que se pueda modificar y perder el historial??
    p=[]
    detalles = DetalleDeItemInspeccion.objects.all()
    if planilla is not None:
        detallesPlanilla = planilla.detalles.all()
        if detallesPlanilla is not None:
            aux=0
            for d in detalles:
                for i in detallesPlanilla:
                    if i.id==d.id:
                        b=[1,d]
                        p.append(b)
                        aux=1
                        break;
                if aux==0:
                    b = [0, d]
                    p.append(b)
                else:
                    aux=0
    items = ItemInspeccion.objects.all()
    categorias = CategoriaInspeccion.objects.all()
    contexto = {"tramite":tramite, "items":items,"detalles":detalles,"categorias":categorias,"movil":movil,"planilla":p}
    return render(request, 'persona/movil/planilla_inspeccion.html', contexto)

def es_movil(request):
    if (request.user_agent.is_mobile):
        return True
    else:
        return False

def inspecciones_realizadas_durante_el_anio(request):
    year=date.today()
    tramites1=Tramite.objects.all()
    tramitesEstado=Estado.objects.filter(timestamp__range=(datetime.date(year.year,01,01), datetime.date(year.year,12,31)), tipo=(6))
    tramites=[]
    for i in range(0, len(tramitesEstado)):
        aux=tramites1.filter(id=tramitesEstado[i].tramite_id).exclude(id__isnull=True)
        tramites.append({"tramite": aux, "fecha": tramitesEstado[i].timestamp})
    return {"tramites":tramites}
